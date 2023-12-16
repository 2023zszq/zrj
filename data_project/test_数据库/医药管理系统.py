import pymysql
import os
import openpyxl
from tkinter import ttk
import tkinter as tk
import tkinter.font as tkFont
from tkinter import *  # 图形界面库
import tkinter.messagebox as messagebox  # 弹窗

class StartPage:
    def __init__(self, parent_window):
        parent_window.destroy()  # 销毁子界面
        self.window = tk.Tk()  # 初始框的声明
        self.window.title('医药管理系统')
        self.window.geometry('800x450')

        # 画布功能添加图片 目前仅支持gif格式的图片
        canvas = tk.Canvas(self.window, height=450, width=800)
        imge_file = tk.PhotoImage(file='test_数据库\首页_800x450.gif')
        imge = canvas.create_image(0, 0, anchor='nw', image=imge_file)
        canvas.pack(side='top')
        Button(self.window, text="管理员登入", font=tkFont.Font(size=16, weight='bold'), command=lambda: AdminPage(self.window),
               width=20,
               height=3,
               fg='black', bg='#C1FFC1', activebackground='black', activeforeground='white').place(x=13, y=355)
        Button(self.window, text="经办人登陆", font=tkFont.Font(size=16, weight='bold'), command=lambda: AgencyPage(self.window),
               width=20,
               height=3, fg='black', bg='#C1FFC1', activebackground='black', activeforeground='white').place(x=280, y=355)
        Button(self.window, text='退出系统', height=3, font=tkFont.Font(size=16, weight='bold'), command=lambda: Exit_operator(self.window), width=20,
               fg='black', bg='#C1FFC1', activebackground='black', activeforeground='white').place(x=540, y=355)

        self.window.mainloop()  # 主消息循环



class AdminPage:
    def __init__(self, parent_window):
        parent_window.destroy()  # 销毁主界面

        self.window = tk.Tk()
        self.window.title('管理员登录')
        self.window.geometry('450x300')

        # welcome image
        canvas = tk.Canvas(self.window, height=300, width=450)
        imge_file = tk.PhotoImage(file='test_数据库\管理员登入_450x300.gif')
        imge = canvas.create_image(0, 0, anchor='nw', image=imge_file)
        canvas.pack(side='top')

        # user information 两个标签文本 x一致 y相差40
        tk.Label(self.window, text='账号').place(x=100, y=150)
        tk.Label(self.window, text='密码').place(x=100, y=190)

        ##############################输入相关信息###################################
        self.var_usr_name = tk.StringVar()
        self.var_usr_name.set('请输入账号')  # 默认用户名 提示用户
        self.var_usr_pwd = tk.StringVar()


        # 输入用户的姓名相关信息
        self.entry_usr_name = tk.Entry(self.window, textvariable=self.var_usr_name)
        self.entry_usr_name.place(x=160, y=150)

        # 输入用户的密码相关信息
        self.entry_usr_pwd = tk.Entry(self.window, textvariable=self.var_usr_pwd, show='*')
        self.entry_usr_pwd.place(x=160, y=190)
        ###################################
        btn_login = tk.Button(self.window, text='登录', command=self.login)
        btn_login.place(x=150, y=230)
        btn_sign_up = tk.Button(self.window, text='返回首页', command=self.back)
        btn_sign_up.place(x=270, y=230)

        self.window.mainloop()

    def login(self):
        # 与数据库进行链接 并判断管理员账号与密码是否正确

        admin_pass = None

        # 数据库操作 查询管理员表
        db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root',
                             password='123456')  # 打开数据库连接
        cursor = db.cursor()  # 使用cursor()方法获取操作游标
        sql = "SELECT * FROM boss_login WHERE boss_id = '%s'" % (self.entry_usr_name.get())  # SQL 查询语句
        try:
            # 执行SQL语句
            cursor.execute(sql)
            # 获取所有记录列表
            results = cursor.fetchall()
            for row in results:
                admin_id = row[0]
                admin_pass = row[1]
                # 打印结果
                print("admin_id=%s,admin pass=%s" % (admin_id, admin_pass))
        except:
            print("Error")
            messagebox.showinfo('警告！', '用户名或密码不正确！')
        db.close()  # 关闭数据库连接

        if self.entry_usr_pwd.get() == admin_pass:
            messagebox.showinfo('提示', '登陆成功！')
            AdminManage(self.window)  # 进入管理员操作界面
        else:
            messagebox.showinfo('警告！', '用户名或密码不正确！')

    def back(self):
        StartPage(self.window)  # 回退至主窗口






class AgencyPage:
    def __init__(self, parent_window):
        parent_window.destroy()  # 销毁主界面
        self.window = tk.Tk()
        self.window.title('经办人登录')
        self.window.geometry('450x300')

        # welcome image
        canvas = tk.Canvas(self.window, height=300, width=450)
        imge_file = tk.PhotoImage(file='test_数据库\经办人登录.gif')
        imge = canvas.create_image(0, 0, anchor='nw', image=imge_file)
        canvas.pack(side='top')

        # user information 两个标签文本 x一致 y相差40
        tk.Label(self.window, text='账号').place(x=100, y=150)
        tk.Label(self.window, text='密码').place(x=100, y=190)

        ##############################输入相关信息###################################
        self.var_agency_name = tk.StringVar()
        self.var_agency_name.set('请输入账号')  # 默认用户名 提示用户
        self.var_agency_pwd = tk.StringVar()

        # 输入用户的姓名相关信息
        self.entry_agency_name = tk.Entry(self.window, textvariable=self.var_agency_name)
        self.entry_agency_name.place(x=160, y=150)

        # 输入用户的密码相关信息
        self.entry_agency_pwd = tk.Entry(self.window, textvariable=self.var_agency_pwd, show='*')
        self.entry_agency_pwd.place(x=160, y=190)
        ###################################
        btn_login = tk.Button(self.window, text='登录', command=self.login)
        btn_login.place(x=150, y=230)
        btn_sign_up = tk.Button(self.window, text='返回首页', command=self.back)
        btn_sign_up.place(x=270, y=230)

        self.window.mainloop()

    def login(self):
        # 与数据库进行链接 并判断管理员账号与密码是否正确
        agency_pass = None

        # 数据库操作 查询管理员表
        db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root',
                             password='123456')  # 打开数据库连接
        cursor = db.cursor()  # 使用cursor()方法获取操作游标
        sql = "SELECT * FROM agency_login WHERE a_id = '%s'" % (self.entry_agency_name.get())  # SQL 查询语句
        try:
            # 执行SQL语句
            cursor.execute(sql)
            # 获取所有记录列表
            results = cursor.fetchall()
            for row in results:
                agency_id = row[0]
                agency_pass = row[1]
                # 打印结果
                print("agency_id=%s,agency_pass=%s" % (agency_id, agency_pass))
        except:
            print("Error")
            messagebox.showinfo('警告！', '用户名或密码不正确！')
        db.close()  # 关闭数据库连接

        if self.entry_agency_pwd.get() == agency_pass:
            messagebox.showinfo('提示', '登陆成功！')
            AgencyManage(self.window)  # 进入经办人操作界面
        else:
            messagebox.showinfo('警告！', '用户名或密码不正确！')

    def back(self):
        StartPage(self.window)  # 回退至主窗口



class AdminManage:#管理员操作界面
    def __init__(self,parent_window):
        self.row = None
        self.col = None
        self.row_info = None
        parent_window.destroy()  # 销毁主界面

        self.window = Tk()  # 初始框的声明
        self.window.title('管理员操作界面')

        self.frame_left_top = tk.Frame(width=300, height=250)
        self.frame_right_top = tk.Frame(width=200, height=300) #200 200
        self.frame_center = tk.Frame(width=750, height=400)
        self.frame_bottom = tk.Frame(width=650, height=50)

        # 定义下方中心列表区域
        self.columns = ("编号", "姓名", "性别", "电话","备注")
        self.tree = ttk.Treeview(self.frame_center, show="headings", height=18, columns=self.columns)
        self.vbar = ttk.Scrollbar(self.frame_center, orient=VERTICAL, command=self.tree.yview)
        # 定义树形结构与滚动条
        self.tree.configure(yscrollcommand=self.vbar.set)

        # 表格的标题
        self.tree.column("编号", width=150, anchor='center')  # 表示列,不显示
        self.tree.column("姓名", width=150, anchor='center')
        self.tree.column("性别", width=100, anchor='center')
        self.tree.column("电话", width=100, anchor='center')
        self.tree.column("备注", width=200, anchor='center')

        self.tree.heading('编号', text="编号", anchor='center')
        self.tree.heading("姓名", text="姓名", anchor='center')
        self.tree.heading("性别", text="性别", anchor='center')
        self.tree.heading("备注", text="备注", anchor='center')

        # 调用方法获取表格内容插入
        self.tree.grid(row=0, column=0, sticky=NSEW)
        self.vbar.grid(row=0, column=1, sticky=NS)

        self.id = []
        self.name = []
        self.gender = []
        self.phone = []
        self.remark = []
        # 打开数据库连接
        db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root', password='123456')
        cursor = db.cursor()  # 使用cursor()方法获取操作游标
        sql = "SELECT * FROM agency"  # SQL 查询语句
        try:
            # 执行SQL语句
            cursor.execute(sql)
            # 获取所有记录列表
            results = cursor.fetchall()
            for row in results:
                self.id.append(row[0])
                self.name.append(row[1])
                self.gender.append(row[2])
                self.phone.append(row[3])
                self.remark.append(row[4])

        except:
            print("Error: unable to fetch data")
            messagebox.showinfo('警告！', '数据库连接失败！')
        db.close()  # 关闭数据库连接

        print("test***********************")
        for i in range(min(len(self.id), len(self.name), len(self.gender), len(self.phone),len(self.remark))):  # 写入数据
            self.tree.insert('', i, values=(self.id[i], self.name[i], self.gender[i], self.phone[i],self.remark[i]))

        for col in self.columns:  # 绑定函数，使表头可排序
            self.tree.heading(col, text=col,
                              command=lambda _col=col: self.tree_sort_column(self.tree, _col, False))

        # 定义顶部区域
        # 定义左上方区域
        self.top_title = Label(self.frame_left_top, text="经办人信息:", font=('Verdana', 20))
        self.top_title.grid(row=0, column=0, columnspan=2, sticky=NSEW, padx=50, pady=10)

        self.left_top_frame = tk.Frame(self.frame_left_top)
        self.var_id = StringVar()  # 声明学号
        self.var_name = StringVar()  # 声明姓名
        self.var_gender = StringVar()  # 声明性别
        self.var_phone = StringVar()  # 声明电话
        self.var_remark = StringVar()  # 声明备注
        # 编号
        self.right_top_id_label = Label(self.frame_left_top, text="编号：", font=('Verdana', 15))
        self.right_top_id_entry = Entry(self.frame_left_top, textvariable=self.var_id, font=('Verdana', 15))
        self.right_top_id_label.grid(row=1, column=0)  # 位置设置
        self.right_top_id_entry.grid(row=1, column=1)
        # 姓名
        self.right_top_name_label = Label(self.frame_left_top, text="姓名：", font=('Verdana', 15))
        self.right_top_name_entry = Entry(self.frame_left_top, textvariable=self.var_name, font=('Verdana', 15))
        self.right_top_name_label.grid(row=2, column=0)  # 位置设置
        self.right_top_name_entry.grid(row=2, column=1)
        # 性别
        self.right_top_gender_label = Label(self.frame_left_top, text="性别：", font=('Verdana', 15))
        self.right_top_gender_entry = Entry(self.frame_left_top, textvariable=self.var_gender,
                                            font=('Verdana', 15))
        self.right_top_gender_label.grid(row=3, column=0)  # 位置设置
        self.right_top_gender_entry.grid(row=3, column=1)
        # 电话
        self.right_top_gender_label = Label(self.frame_left_top, text="电话：", font=('Verdana', 15))
        self.right_top_gender_entry = Entry(self.frame_left_top, textvariable=self.var_phone,
                                            font=('Verdana', 15))
        self.right_top_gender_label.grid(row=4, column=0)  # 位置设置
        self.right_top_gender_entry.grid(row=4, column=1)
        
        #备注
        self.right_top_gender_label = Label(self.frame_left_top, text="备注：", font=('Verdana', 15))
        self.right_top_gender_entry = Entry(self.frame_left_top, textvariable=self.var_remark,
                                            font=('Verdana', 15))
        self.right_top_gender_label.grid(row=5, column=0)  # 位置设置
        self.right_top_gender_entry.grid(row=5, column=1)

        # 定义右上方区域
        self.right_top_title = Label(self.frame_right_top, text="操作：", font=('Verdana', 20))

        self.tree.bind('<Button-1>', self.click)  # 左键获取位置
        self.right_top_button1 = ttk.Button(self.frame_right_top, text='新建经办人', width=20, command=self.new_row)
        self.right_top_button2 = ttk.Button(self.frame_right_top, text='更新经办人信息', width=20,
                                            command=self.updata_row)
        self.right_top_button3 = ttk.Button(self.frame_right_top, text='删除选中经办人信息', width=20,
                                            command=self.del_row)
        self.right_top_button4 = ttk.Button(self.frame_right_top, text='经办人数据报表', width=20,
                                            command=self.operator_data_report)
        self.right_top_button5 = ttk.Button(self.frame_right_top, text='返回', width=20,
                                            command=self.management_interface_return)

        # 位置设置
        self.right_top_title.grid(row=1, column=0, pady=10)
        self.right_top_button1.grid(row=2, column=0, padx=20, pady=10)
        self.right_top_button2.grid(row=3, column=0, padx=20, pady=10)
        self.right_top_button3.grid(row=4, column=0, padx=20, pady=10)
        self.right_top_button4.grid(row=5, column=0, padx=20, pady=10)
        self.right_top_button5.grid(row=6, column=0, padx=20, pady=10)

        # 整体区域定位
        self.frame_left_top.grid(row=0, column=0, padx=2, pady=5)
        self.frame_right_top.grid(row=0, column=1, padx=30, pady=30)
        self.frame_center.grid(row=1, column=0, columnspan=2, padx=4, pady=5)
        self.frame_bottom.grid(row=2, column=0, columnspan=2)

        self.frame_left_top.grid_propagate(False)
        self.frame_right_top.grid_propagate(False)
        self.frame_center.grid_propagate(False)
        self.frame_bottom.grid_propagate(False)

        self.frame_left_top.tkraise()  # 开始显示主菜单
        self.frame_right_top.tkraise()  # 开始显示主菜单
        self.frame_center.tkraise()  # 开始显示主菜单
        self.frame_bottom.tkraise()  # 开始显示主菜单

        self.window.protocol("WM_DELETE_WINDOW", self.back)  # 捕捉右上角关闭点击，回到首页

    def management_interface_return(self):
        StartPage(self.window)
    
    def operator_data_report(self):
        
        host = "localhost"
        username = "root"
        passwd = "123456"
        db = "medicine management system"
        port = 3306
        charset = "utf8"
        excel_name = "经办人信息.xlsx"
        table_name='agency'
        sheet=[]
        conn = pymysql.connect(host=host, user=username,
                        passwd=passwd, db=db,
                        port=port, charset=charset)
        cur = conn.cursor()
        sql = "select * from %s;" % table_name
        cur.execute(sql)
        sql_result = cur.fetchall()
        cur.close()
        conn.close()
        # 写 Excel
        book = openpyxl.Workbook()
        sheet = book.active
        # fff = [filed[0] for filed in cur.description]  # 获取表头信息
        fff=['经办人编号','经办人姓名','性别','电话号码','备注']
        sheet.append(fff) # type: ignore
        
        for i in sql_result:
            sheet.append(i) # type: ignore
        
        try:
            book.save("%s" % excel_name)
        except:
            messagebox.showinfo('警告！', '数据报表失败!')
        path_=os.getcwd()
        messagebox.showinfo('提示！', f'经办人信息数据报表创建成功!保存路径为:{path_}')
        
            


    def back(self):
        StartPage(self.window)  # 显示主窗口 销毁本窗口

    def click(self,event):#click点击 将列表的数据显示在输入框内
        self.col = self.tree.identify_column(event.x)  # 列
        self.row = self.tree.identify_row(event.y)  # 行

        print(self.col)
        print(self.row)
        self.row_info = self.tree.item(self.row, "values")
        self.var_id.set(self.row_info[0])
        self.var_name.set(self.row_info[1])
        self.var_gender.set(self.row_info[2])
        self.var_phone.set(self.row_info[3])
        self.var_remark.set(self.row_info[4])
        self.right_top_id_entry = Entry(self.frame_left_top, state='disabled', textvariable=self.var_id,
                                        font=('Verdana', 15))
        print()



    def tree_sort_column(self, tv, col, reverse):  # Treeview、列名、排列方式
        l = [(tv.set(k, col), k) for k in tv.get_children('')]
        l.sort(reverse=reverse)  # 排序方式
        # rearrange items in sorted positions
        for index, (val, k) in enumerate(l):  # 根据排序后索引移动
            tv.move(k, '', index)
        tv.heading(col, command=lambda: self.tree_sort_column(tv, col, not reverse))  # 重写标题，使之成为再点倒序的标题




    def new_row(self):#创建新信息

        # print(self.var_id.get())
        # print(self.id)
        if str(self.var_id.get()) in self.id:
            messagebox.showinfo('警告！', '该经办人已存在！')
        else:
            if self.var_id.get() != '' and self.var_name.get() != '' and self.var_gender.get() != '' and self.var_phone.get() != '':
                # 打开数据库连接
                db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root', password='123456')
                cursor = db.cursor()  # 使用cursor()方法获取操作游标
                sql = "INSERT INTO agency(ano, aname, asex, aphone,aremark) \
        				       VALUES ('%s', '%s', '%s', '%s', '%s')" % \
                      (self.var_id.get(), self.var_name.get(), self.var_gender.get(), self.var_phone.get(),self.var_remark.get())  # SQL 插入语句

                try:
                    cursor.execute(sql)  # 执行sql语句
                    db.commit()  # 提交到数据库执行
                except:
                    db.rollback()  # 发生错误时回滚
                    messagebox.showinfo('警告！', '数据库连接失败！')
                db.close()  # 关闭数据库连接

                self.id.append(self.var_id.get())
                self.name.append(self.var_name.get())
                self.gender.append(self.var_gender.get())
                self.phone.append(self.var_phone.get())
                self.remark.append(self.var_remark.get())
                self.tree.insert('', len(self.id) - 1, values=(
                    self.id[len(self.id) - 1], self.name[len(self.id) - 1], self.gender[len(self.id) - 1],
                    self.phone[len(self.id) - 1],self.remark[len(self.id) - 1])),
                self.tree.update()

                messagebox.showinfo('提示！', '新建经办人成功！')
            else:
                messagebox.showinfo('警告！', '请填写经办人数据')






    def updata_row(self):
        res = messagebox.askokcancel('警告！', '是否更新所填数据？')
        if res:
            if self.var_id.get() == self.row_info[0]:  # 如果所填学号 与 所选学号一致
                # 打开数据库连接
                db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root', password='123456')
                cursor = db.cursor()  # 使用cursor()方法获取操作游标
                sql_update = "UPDATE agency SET aname = '%s', asex = '%s', aphone = '%s' ,aremark = '%s'\
        				 WHERE ano = '%s'" % (
                    self.var_name.get(), self.var_gender.get(), self.var_phone.get(), self.var_remark.get(),self.var_id.get())  # SQL 插入语句
                try:
                    cursor.execute(sql_update)  # 执行sql语句
                    db.commit()  # 提交到数据库执行
                    messagebox.showinfo('提示！', '更新成功！')
                except:
                    db.rollback()  # 发生错误时回滚
                    messagebox.showerror('警告！', '更新失败，数据库连接失败！')
                db.close()  # 关闭数据库连接

                id_index = self.id.index(self.row_info[0])
                self.name[id_index] = self.var_name.get()
                self.gender[id_index] = self.var_gender.get()
                self.phone[id_index] = self.var_phone.get()
                self.remark[id_index]=self.var_remark.get()

                self.tree.item(self.tree.selection()[0], values=(
                    self.var_id.get(), self.var_name.get(), self.var_gender.get(),
                    self.var_phone.get(),self.var_remark.get()))  # 修改对于行信息
            else:
                messagebox.showinfo('警告！', '不能修改经办人编号！')

        # db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root',
        #                          password='123456')
        # cursor = db.cursor()  # 使用cursor()方法获取操作游标
        # sql_update = "UPDATE agency SET aname = '%s', asex = '%s', aphone = '%s' \
        #            				 WHERE ano = '%s'" % (
        #     self.var_name.get(), self.var_gender.get(), self.var_phone.get(), self.var_id.get())  # SQL 插入语句
        # cursor.execute(sql_update)  # 执行sql语句
        # db.commit()  # 提交到数据库执行
        #
        # db.close()
    def del_row(self):
        res = messagebox.askokcancel('警告！', '是否删除所选数据？')
        if res:
            # 打开数据库连接
            db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root',
                                 password='123456')
            cursor = db.cursor()  # 使用cursor()方法获取操作游标
            sql_delete = "DELETE FROM agency WHERE ano = '%s'" % (self.row_info[0])  # SQL 插入语句
            #sql_deletes = "DELETE  FROM stu_login_k WHERE stu_id = '%s'" % (self.row_info[0])  # SQL 插入语句
            try:
                cursor.execute(sql_delete)  # 执行sql语句
                #cursor.execute(sql_deletes)
                db.commit()  # 提交到数据库执行
                messagebox.showinfo('提示！', '删除成功！')
            except:
                db.rollback()  # 发生错误时回滚
                messagebox.showinfo('警告！', '删除失败，数据库连接失败！')
            db.close()  # 关闭数据库连接

            id_index = self.id.index(self.row_info[0])
            print(id_index)
            del self.id[id_index]
            del self.name[id_index]
            del self.gender[id_index]
            del self.phone[id_index]
            del self.remark[id_index]
            print(self.id)
            self.tree.delete(self.tree.selection()[0])  # 删除所选行
            print(self.tree.get_children())





class AgencyManage:  # 经办人操作界面
    def __init__(self, parent_window):
        parent_window.destroy()  # 销毁子界面
        self.window = tk.Tk()  # 初始框的声明
        self.window.title('经办人操作界面')
        self.window.geometry('800x449')

        # 画布功能添加图片 目前仅支持gif格式的图片
        canvas = tk.Canvas(self.window, height=450, width=800)
        imge_file = tk.PhotoImage(file='test_数据库\经办人操作界面.gif')
        imge = canvas.create_image(0, 0, anchor='nw', image=imge_file)
        canvas.pack(side='top')
        # lamada 相当于定义函数
        Button(self.window, text="录入患者信息", font=tkFont.Font(size=16,weight='bold'),
               command=lambda: Customer_data_entry(self.window), width=17,
               height=3,
               fg='black', bg='#B0E0E6', activebackground='black', activeforeground='dimgray').place(x=582,y=8)

        Button(self.window, text="录入药品信息", font=tkFont.Font(size=16,weight='bold'),
               command=lambda: Drug_information_entry(self.window), width=17,
               height=3, fg='black', bg='#B0E0E6', activebackground='black', activeforeground='white').place(x=582,y=96)

        Button(self.window, text="查看患者信息", font=tkFont.Font(size=16,weight='bold'),
               command=lambda: Customer_select(self.window), width=17, height=3, fg='black',
               bg='#B0E0E6', activebackground='black', activeforeground='white').place(x=582,y=182)

        Button(self.window, text='查看药品信息', height=3, font=tkFont.Font(size=16,weight='bold'),
               command=lambda: Check_drug_information(self.window), width=17,
               fg='black', bg='#B0E0E6', activebackground='black', activeforeground='white').place(x=582,y=270)

        Button(self.window, text='退出经办人操作系统', height=3, font=tkFont.Font(size=16,weight='bold'),
               command=self.exit_operator_operating_system, width=17,
               fg='black', bg='#B0E0E6', activebackground='black', activeforeground='white').place(x=582,y=358)
        self.window.mainloop()  # 主消息循环
    def exit_operator_operating_system(self):
        StartPage(self.window)


class Customer_data_entry:
    def __init__(self, parent_window):
        total_ano=[]
        self.total_ano = []
        self.row = None
        self.col = None
        self.row_info = None
        parent_window.destroy()  # 销毁主界面

        self.window = Tk()  # 初始框的声明
        self.window.title('患者信息')

        self.frame_left_top = tk.Frame(width=500, height=400)
        self.frame_right_top = tk.Frame(width=400, height=300)
        self.frame_center = tk.Frame(width=1200, height=300)
        self.frame_bottom = tk.Frame(width=950, height=250)

        # 定义下方中心列表区域
        self.columns = ("编号", "姓名", "性别", "年龄","住址","电话","症状","已购药品","经办人","录入日期","备注")
        self.tree = ttk.Treeview(self.frame_center, show="headings", height=30, columns=self.columns)#18
        self.vbar = ttk.Scrollbar(self.frame_center, orient=VERTICAL, command=self.tree.yview)
        # 定义树形结构与滚动条
        self.tree.configure(yscrollcommand=self.vbar.set)

        # 表格的标题
        self.tree.column("编号", width=150, anchor='center')  # 表示列,不显示
        self.tree.column("姓名", width=150, anchor='center')
        self.tree.column("性别", width=100, anchor='center')
        self.tree.column("年龄", width=100, anchor='center')
        self.tree.column("住址", width=100, anchor='center')
        self.tree.column("电话", width=100, anchor='center')
        self.tree.column("症状", width=100, anchor='center')
        self.tree.column("已购药品", width=100, anchor='center')
        self.tree.column("经办人", width=100, anchor='center')
        self.tree.column("录入日期", width=100, anchor='center')
        self.tree.column("备注", width=100, anchor='center')


        # 调用方法获取表格内容插入
        self.tree.grid(row=0, column=0, sticky=NSEW)
        self.vbar.grid(row=0, column=1, sticky=NS)

        self.id = []
        self.name = []
        self.gender = []
        self.age = []
        self.address=[]
        self.phone=[]
        self.symptom=[]
        self.mno=[]
        self.ano=[]
        self.date=[]
        self.remark=[]
        # 打开数据库连接
        db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root',
                             password='123456')
        cursor = db.cursor()  # 使用cursor()方法获取操作游标
        sql = "SELECT * FROM customer_data_copy1"  # SQL 查询语句 customer_data_copy1 customer_d
        try:
            # 执行SQL语句
            cursor.execute(sql)
            # 获取所有记录列表
            results = cursor.fetchall()
            for row in results:
                self.id.append(row[0])
                self.name.append(row[1])
                self.gender.append(row[2])
                self.age.append(row[3])
                self.address.append(row[4])
                self.phone.append(row[5])
                self.symptom.append(row[6])
                self.mno.append(row[7])
                self.ano.append(row[8])
                self.date.append(row[9])
                self.remark.append(row[10])


        except:
            print("Error: unable to fetch data")
            messagebox.showinfo('警告！', '数据库连接失败！')
        db.close()  # 关闭数据库连接

        print("test***********************")
        for i in range(min(len(self.id), len(self.name), len(self.gender), len(self.phone))):  # 写入数据
            self.tree.insert('', i, values=(self.id[i], self.name[i], self.gender[i], self.age[i],self.address[i],self.phone[i],self.symptom[i],
                                            self.mno[i],self.ano[i],self.date[i],self.remark[i]))
        for col in self.columns:  # 绑定函数，使表头可排序
            self.tree.heading(col, text=col,
                              command=lambda _col=col: self.tree_sort_column(self.tree, _col, False))

        # 定义顶部区域
        # 定义左上方区域
        self.top_title = Label(self.frame_left_top, text="患者信息:", font=('Verdana', 20))
        self.top_title.grid(row=0, column=0, columnspan=2, sticky=NSEW, padx=50, pady=10)#50 ,10

        self.left_top_frame = tk.Frame(self.frame_left_top)
        self.var_id = StringVar()  # 声明学号
        self.var_name = StringVar()  # 声明姓名
        self.var_gender = StringVar()  # 声明性别
        self.var_age = StringVar()  # 声明年龄
        self.var_address=StringVar()  # 声明地址
        self.var_phone=StringVar()      # 声明电话
        self.var_symptom=StringVar()    # 声明症状
        self.var_mno=StringVar()    # 声明药品
        self.var_ano=StringVar()    # 声明经办人
        self.var_date=StringVar()   # 声明日期
        self.var_remark=StringVar()     # 声明备注





        # 编号
        self.right_top_id_label = Label(self.frame_left_top, text="编号：", font=('Verdana', 15))
        self.right_top_id_entry = Entry(self.frame_left_top, textvariable=self.var_id, font=('Verdana', 15))
        self.right_top_id_label.grid(row=1, column=0)  # 位置设置
        self.right_top_id_entry.grid(row=1, column=1)
        # 姓名
        self.right_top_name_label = Label(self.frame_left_top, text="姓名：", font=('Verdana', 15))
        self.right_top_name_entry = Entry(self.frame_left_top, textvariable=self.var_name, font=('Verdana', 15))
        self.right_top_name_label.grid(row=2, column=0)  # 位置设置
        self.right_top_name_entry.grid(row=2, column=1)
        # 性别
        self.right_top_gender_label = Label(self.frame_left_top, text="性别：", font=('Verdana', 15))
        self.right_top_gender_entry = Entry(self.frame_left_top, textvariable=self.var_gender,
                                            font=('Verdana', 15))
        self.right_top_gender_label.grid(row=3, column=0)  # 位置设置
        self.right_top_gender_entry.grid(row=3, column=1)
        # 年龄
        self.right_top_gender_label = Label(self.frame_left_top, text="年龄：", font=('Verdana', 15))
        self.right_top_gender_entry = Entry(self.frame_left_top, textvariable=self.var_age,
                                            font=('Verdana', 15))
        self.right_top_gender_label.grid(row=4, column=0)  # 位置设置
        self.right_top_gender_entry.grid(row=4, column=1)

        # 住址
        self.right_top_gender_label = Label(self.frame_left_top, text="住址：", font=('Verdana', 15))
        self.right_top_gender_entry = Entry(self.frame_left_top, textvariable=self.var_address,
                                            font=('Verdana', 15))
        self.right_top_gender_label.grid(row=5, column=0)  # 位置设置
        self.right_top_gender_entry.grid(row=5, column=1)

        # 电话
        self.right_top_gender_label = Label(self.frame_left_top, text="电话：", font=('Verdana', 15))
        self.right_top_gender_entry = Entry(self.frame_left_top, textvariable=self.var_phone,
                                            font=('Verdana', 15))
        self.right_top_gender_label.grid(row=6, column=0)  # 位置设置
        self.right_top_gender_entry.grid(row=6, column=1)

        # 症状
        self.right_top_gender_label = Label(self.frame_left_top, text="症状：", font=('Verdana', 15))
        self.right_top_gender_entry = Entry(self.frame_left_top, textvariable=self.var_symptom,
                                            font=('Verdana', 15))
        self.right_top_gender_label.grid(row=7, column=0)  # 位置设置
        self.right_top_gender_entry.grid(row=7, column=1)

        # 已购药品
        self.right_top_gender_label = Label(self.frame_left_top, text="备注：", font=('Verdana', 15))
        self.right_top_gender_entry = Entry(self.frame_left_top, textvariable=self.var_remark,
                                            font=('Verdana', 15))
        self.right_top_gender_label.grid(row=8, column=0)  # 位置设置
        self.right_top_gender_entry.grid(row=8, column=1)

        # 经办人
        self.right_top_gender_label = Label(self.frame_left_top, text="经办人：", font=('Verdana', 15))
        self.right_top_gender_entry = Entry(self.frame_left_top, textvariable=self.var_ano,
                                            font=('Verdana', 15))
        self.right_top_gender_label.grid(row=9, column=0)  # 位置设置
        self.right_top_gender_entry.grid(row=9, column=1)

        # 录入日期
        self.right_top_gender_label = Label(self.frame_left_top, text="录入日期：", font=('Verdana', 15))
        self.right_top_gender_entry = Entry(self.frame_left_top, textvariable=self.var_date,
                                            font=('Verdana', 15))
        self.right_top_gender_label.grid(row=10, column=0)  # 位置设置
        self.right_top_gender_entry.grid(row=10, column=1)

        # 备注
        self.right_top_gender_label = Label(self.frame_left_top, text="已购药品：", font=('Verdana', 15))
        # self.right_top_gender_entry = Entry(self.frame_left_top, textvariable=self.var_mno,
        #                                     font=('Verdana', 15))
        self.right_top_gender_label.grid(row=11, column=0)  # 位置设置
        # self.right_top_gender_entry.grid(row=11, column=1)
        
        self.right_top_button5 = ttk.Button(self.frame_left_top, text='选择药品', width=20,
                                            command=self.select_medication)
        self.right_top_button5.grid(row=11, column=1)
        
        #选择药品显示
        self.select_drug_display = tk.StringVar()
        self.entry_select_drug_display=tk.Entry(self.frame_left_top, textvariable=self.select_drug_display,width=50)
        self.entry_select_drug_display.grid(row=11,column=2)
        
        

        





        # 定义右上方区域
        self.right_top_title = Label(self.frame_right_top, text="操作：", font=('Verdana', 20))

        self.tree.bind('<Button-1>', self.click)  # 左键获取位置
        self.right_top_button1 = ttk.Button(self.frame_right_top, text='新建患者信息', width=20, command=self.new_row)
        self.right_top_button2 = ttk.Button(self.frame_right_top, text='更新患者信息', width=20,
                                            command=self.updata_row)
        self.right_top_button3 = ttk.Button(self.frame_right_top, text='删除选中患者信息', width=20,
                                            command=self.del_row)
        self.right_top_button4 = ttk.Button(self.frame_right_top, text='返回', width=20,
                                            command=self.back)

        # 位置设置
        self.right_top_title.grid(row=1, column=0, pady=10)
        self.right_top_button1.grid(row=2, column=0, padx=20, pady=10)
        self.right_top_button2.grid(row=3, column=0, padx=20, pady=10)
        self.right_top_button3.grid(row=4, column=0, padx=20, pady=10)
        self.right_top_button4.grid(row=5,column=0,padx=20, pady=10)
        # 整体区域定位

        self.frame_left_top.grid(row=0, column=0, padx=2, pady=5)
        self.frame_right_top.grid(row=0, column=1, padx=30, pady=30)
        self.frame_center.grid(row=1, column=0, columnspan=2, padx=4, pady=5)
        self.frame_bottom.grid(row=2, column=0, columnspan=2)

        self.frame_left_top.grid_propagate(False)
        self.frame_right_top.grid_propagate(False)
        self.frame_center.grid_propagate(False)
        self.frame_bottom.grid_propagate(False)

        self.frame_left_top.tkraise()  # 开始显示主菜单
        self.frame_right_top.tkraise()  # 开始显示主菜单
        self.frame_center.tkraise()  # 开始显示主菜单
        self.frame_bottom.tkraise()  # 开始显示主菜单

        self.window.protocol("WM_DELETE_WINDOW", self.back)  # 捕捉右上角关闭点击，回到首页

    def tree_sort_column(self, tv, col, reverse):  # Treeview、列名、排列方式
        l = [(tv.set(k, col), k) for k in tv.get_children('')]
        l.sort(reverse=reverse)  # 排序方式
        # rearrange items in sorted positions
        for index, (val, k) in enumerate(l):  # 根据排序后索引移动
            tv.move(k, '', index)
        tv.heading(col, command=lambda: self.tree_sort_column(tv, col, not reverse))  # 重写标题，使之成为再点倒序的标题

    def back(self):
        AgencyManage(self.window)  #回退至经办人操作窗口



    def click(self,event):#click点击 将列表的数据显示在输入框内
        self.col = self.tree.identify_column(event.x)  # 列
        self.row = self.tree.identify_row(event.y)  # 行

        print(self.col)
        print(self.row)
        self.row_info = self.tree.item(self.row, "values")
        self.var_id.set(self.row_info[0])
        self.var_name.set(self.row_info[1])
        self.var_gender.set(self.row_info[2])
        self.var_age.set(self.row_info[3])
        self.var_address.set(self.row_info[4])
        self.var_phone.set(self.row_info[5])
        self.var_symptom.set(self.row_info[6])
        self.select_drug_display.set(self.row_info[7])
        self.var_ano.set(self.row_info[8])
        self.var_date.set(self.row_info[9])
        self.var_remark.set(self.row_info[10])
        self.right_top_id_entry = Entry(self.frame_left_top, state='disabled', textvariable=self.var_id,
                                        font=('Verdana', 15))
        print()


    #添加顾客信息
    def new_row(self):
        flag=True
        db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root',
                             password='123456')
        cursor = db.cursor()  # 使用cursor()方法获取操作游标
        sql = "SELECT ano FROM agency"  # SQL 查询语句
        try:
            # 执行SQL语句
            cursor.execute(sql)
            # 获取所有记录列表
            results = cursor.fetchall()
            for row in results:
                self.total_ano.append(row[0])

        except:
            print("Error: unable to fetch data")
            messagebox.showinfo('警告！', '数据库连接失败！')
        db.close()  # 关闭数据库连接

        print(self.total_ano)


        if str(self.var_id.get()) in self.id:
            messagebox.showinfo('警告！', '该患者编号信息已存在！')
        else:
            #判断输入信息是否为空
            if self.var_id.get() != '' and self.var_name.get() != '' and self.var_gender.get() != '' and self.var_phone.get() != '':
                if self.var_ano.get() in self.total_ano:
                    # 打开数据库连接
                    db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root', password='123456')
                    cursor = db.cursor()  # 使用cursor()方法获取操作游标
                # sql = "INSERT INTO customer_d(no_c, name, csex, cage, caddress, cphone , csymptom,  mno, ano, cdate, cremark) \
        		# 		       VALUES ('%s', '%s', '%s', '%s','%s', '%s', '%s', '%s','%s', '%s', '%s')" % \
                #       (self.var_id.get(), self.var_name.get(), self.var_gender.get(), self.var_age.get(),self.var_address.get(),
                #        self.var_phone.get(), self.var_symptom.get(), self.var_mno.get(),self.var_ano.get(), self.var_date.get(), self.var_remark.get())  # SQL 插入语句

                # sql = "INSERT INTO Customer_data(cno) \
                # 		       VALUES ('%s')" % \
                #       (self.var_id.get())  # SQL 插入语句

                # sql="""INSERT INTO agency(ano, aname, asex, aphone) \
        		# 		       VALUES ('%s', '%s', '%s', '%s')""" % \
                #       (self.var_id.get(), self.var_name.get(), self.var_gender.get(), self.var_phone.get()

                    sql = "INSERT INTO customer_data_copy1(cno, cname, csex, cage, caddress, cphone , csymptom,  mno, ano, cdate, cremark) \
                                   VALUES ('%s', '%s', '%s', '%s','%s', '%s', '%s', '%s','%s', '%s', '%s')" % \
                          (self.var_id.get(), self.var_name.get(), self.var_gender.get(), self.var_age.get(),self.var_address.get(),
                           self.var_phone.get(), self.var_symptom.get(), self.select_drug_display.get(),self.var_ano.get(), self.var_date.get(), self.var_remark.get())  # SQL 插入语句

                    try:
                        cursor.execute(sql)  # 执行sql语句
                        db.commit()  # 提交到数据库执行
                    except:
                        db.rollback()  # 发生错误时回滚
                        messagebox.showinfo('警告！', '数据库连接失败！')
                    db.close()  # 关闭数据库连接
                else:
                    flag = False
                    messagebox.showinfo('警告！', '该经办人编号不存在')


                if flag:
                    self.id.append(self.var_id.get())
                    self.name.append(self.var_name.get())
                    self.gender.append(self.var_gender.get())
                    self.age.append(self.var_age.get())
                    self.address.append(self.var_address.get())
                    self.phone.append(self.var_phone.get())
                    self.symptom.append(self.var_symptom.get())
                    self.mno.append(self.select_drug_display.get())
                    self.ano.append(self.var_ano.get())
                    self.date.append(self.var_date.get())
                    self.remark.append(self.var_remark.get())
                    self.tree.insert('', len(self.id) - 1, values=(
                        self.id[len(self.id) - 1], self.name[len(self.id) - 1], self.gender[len(self.id) - 1],self.age[len(self.id) - 1],
                        self.address[len(self.id) - 1],
                        self.phone[len(self.id) - 1],
                        self.symptom[len(self.id) - 1],
                        self.mno[len(self.id) - 1],
                        self.ano[len(self.id) - 1],
                        self.date[len(self.id) - 1],
                        self.remark[len(self.id) - 1],
                    )),

                    self.tree.update()
                    messagebox.showinfo('提示！', '新建患者信息成功！')
                    self.tree.update()
            else:
                messagebox.showinfo('警告！', '请填写患者数据')




    def updata_row(self):
        flag=True
        res = messagebox.askokcancel('警告！', '是否更新所填数据？')#按钮返回true或flase
        db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root',
                             password='123456')
        cursor = db.cursor()  # 使用cursor()方法获取操作游标
        sql = "SELECT ano FROM agency"  # SQL 查询语句
        try:
            # 执行SQL语句
            cursor.execute(sql)
            # 获取所有记录列表
            results = cursor.fetchall()
            for row in results:
                self.total_ano.append(row[0])

        except:
            print("Error: unable to fetch data")
            messagebox.showinfo('警告！', '数据库连接失败！')
        db.close()  # 关闭数据库连接
        print()
        if res:
            if self.var_id.get() == self.row_info[0]:  # 如果所填学号 与 所选学号一致
                # 打开数据库连接
                if self.var_ano.get() in self.total_ano:
                    db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root',
                                         password='123456')
                    cursor = db.cursor()  # 使用cursor()方法获取操作游标
                    sql_update = "UPDATE customer_data_copy1 SET cname = '%s', csex = '%s', cage = '%s' ,caddress= '%s'  ,cphone= '%s',csymptom= '%s',mno= '%s',ano= '%s',cdate= '%s',cremark= '%s'\
                                     WHERE cno = '%s'"  %(
                        self.var_name.get(), self.var_gender.get(), self.var_age.get(), self.var_address.get(),
                        self.var_phone.get(), self.var_symptom.get(), self.select_drug_display.get(), self.var_ano.get(),
                        self.var_date.get(), self.var_remark.get(),self.var_id.get())  # SQL 插入语句
                    try:
                        cursor.execute(sql_update)  # 执行sql语句
                        db.commit()  # 提交到数据库执行 
                        messagebox.showinfo('提示！', '更新成功！')
                    except:
                        flag=False
                        db.rollback()  # 发生错误时回滚
                        messagebox.showerror('警告！', '更新失败，数据库连接失败！')
                        db.close()  # 关闭数据库连接
                    if flag:
                        id_index = self.id.index(self.row_info[0])
                        self.name[id_index] = self.var_name.get()
                        self.gender[id_index] = self.var_gender.get()
                        self.phone[id_index] = self.var_phone.get()
                        self.age[id_index]=self.var_age.get()
                        self.address[id_index] =  self.var_address.get()
                        self.phone[id_index] = self.var_phone.get()
                        self.symptom[id_index] = self.var_symptom.get()
                        self.mno[id_index] = self.select_drug_display.get()
                        self.ano[id_index] = self.var_ano.get()
                        self.date[id_index] = self.var_date.get()
                        self.remark[id_index] = self.var_remark.get()


                        self.tree.item(self.tree.selection()[0], values=(
                            self.var_id.get(), self.var_name.get(), self.var_gender.get()
                            ,self.var_age.get(),self.var_address.get(),self.var_phone.get(),self.var_symptom.get()
                        ,self.select_drug_display.get(),self.var_ano.get(),self.var_date.get(),self.var_remark.get()))  # 修改对于行信息
                else:
                    messagebox.showinfo('警告！', '该经办人编号不存在')

            else:
                messagebox.showinfo('警告！', '不能修改患者编号！')

    def del_row(self):
        res = messagebox.askokcancel('警告！', '是否删除所选数据？')
        if res:
            # 打开数据库连接
            db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root',
                                 password='123456')
            cursor = db.cursor()  # 使用cursor()方法获取操作游标
            sql_delete = "DELETE FROM customer_data_copy1 WHERE cno = '%s'" % (self.row_info[0])  # SQL 插入语句
            try:
                cursor.execute(sql_delete)  # 执行sql语句
                db.commit()  # 提交到数据库执行
                messagebox.showinfo('提示！', '删除成功！')
            except:
                db.rollback()  # 发生错误时回滚
                messagebox.showinfo('警告！', '删除失败，数据库连接失败！')
            db.close()  # 关闭数据库连接

            id_index = self.id.index(self.row_info[0])

            del self.id[id_index]
            del self.name[id_index]
            del self.gender[id_index]
            del self.phone[id_index]

            self.tree.delete(self.tree.selection()[0])  # 删除所选行
            
            
            
            
    #选择药品
    def select_medication(self):
        self.window_Check_drug_information = tk.Tk()  # 初始框的声明
        self.window_Check_drug_information.title('查看药品信息')
        self.window_Check_drug_information.geometry('1000x450')

        # 画布功能添加图片 目前仅支持gif格式的图片  test_数据库\222222.gif
        # canvas = tk.Canvas(self.window_Check_drug_information, height=200, width=500)
        # imge_file = tk.PhotoImage(file='test_数据库\药品信息_500x281.gif')
        # imge = canvas.create_image(0, 0, anchor='nw', image=imge_file)
        # canvas.grid(row=0, column=0,padx=250)
        # lamada 相当于定义函数

        self.frame_center_Check_drug_information = tk.Frame(self.window_Check_drug_information,width=950, height=300)

        self.columns_Check_drug_information = (
           "编号", "名称", "服用方法", "功效",'库存')
        self.tree = ttk.Treeview(self.frame_center_Check_drug_information, show="headings", height=10, columns=self.columns_Check_drug_information)  # 18
        self.vbar = ttk.Scrollbar(self.frame_center_Check_drug_information, orient=VERTICAL, command=self.tree.yview)
        # 定义树形结构与滚动条
        self.tree.configure(yscrollcommand=self.vbar.set)

        self.tree.column("编号", width=150, minwidth=100, anchor='center')  # 表示列,不显示
        self.tree.column("名称", width=150, minwidth=100, anchor='center')
        self.tree.column("服用方法", width=100, minwidth=100, anchor='center')
        self.tree.column("功效", width=450, minwidth=100, anchor='center')
        self.tree.column("库存", width=80,minwidth=100, anchor='center')

        self.tree.heading('编号', text="编号", anchor='center')
        self.tree.heading("名称", text="名称", anchor='center')
        self.tree.heading("服用方法", text="服用方法", anchor='center')
        self.tree.heading("功效", text="功效", anchor='center')
        self.tree.heading("库存", text="库存", anchor='center')

        self.tree.grid(row=1, column=0, sticky=NSEW)
        self.vbar.grid(row=1, column=1, sticky=NS)
        
        self.top_label = Label(self.frame_center_Check_drug_information, text="药品信息:", font=('Verdana', 20))
        self.top_label.grid(row=0, column=0, columnspan=2, sticky=NSEW, padx=50, pady=10) 
        
        self.middle = Label(self.frame_center_Check_drug_information, text="已选药品:", font=('Verdana', 20))
        self.middle.grid(row=2, column=0, columnspan=2, sticky=NSEW, padx=50, pady=10) 

        #开始查询框的变量
        self.id_Check_drug_information = []
        self.name_Check_drug_information = []
        self.mode_Check_drug_information = []
        self.mefficacy_Check_drug_information = []
        self.mid_Check_drug_information=[]
        
        #选择药品数据信息
        self.list_drug_information=[]


        #药品编号查询变量：
        self.select_id_mefficacy_Check_drug_information=[]
        self.select_id_mode_Check_drug_information=[]
        self.select_id_name_Check_drug_information=[]
        self.select_id_id_Check_drug_information=[]
        self.select_id_mid_Check_drug_information=[]
        #数据库查询
        db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root',
                             password='123456')
        cursor = db.cursor()  # 使用cursor()方法获取操作游标
        sql = "SELECT * FROM medicine"  # SQL 查询语句 customer_data_copy1 customer_d
        try:
            # 执行SQL语句
            cursor.execute(sql)
            # 获取所有记录列表
            results = cursor.fetchall()
            for row in results:
                self.id_Check_drug_information.append(row[0])
                self.name_Check_drug_information.append(row[1])
                self.mode_Check_drug_information.append(row[2])
                self.mefficacy_Check_drug_information.append(row[3])
                self.mid_Check_drug_information.append(row[4])


        except:
            print("Error: unable to fetch data")
            messagebox.showinfo('警告！', '数据库连接失败！')
        db.close()  # 关闭数据库连接

        for i in range(min(len(self.id_Check_drug_information), len(self.name_Check_drug_information), len(self.mode_Check_drug_information), len(self.mefficacy_Check_drug_information),len(self.mid_Check_drug_information))):  # 写入数据
            self.tree.insert('', i, values=(self.id_Check_drug_information[i], self.name_Check_drug_information[i], self.mode_Check_drug_information[i], self.mefficacy_Check_drug_information[i],self.mid_Check_drug_information[i],
                                            ))

        #药品信息框
        self.frame_center_Check_drug_information.grid(row=1, column=0, columnspan=3, padx=0, pady=0)  # , padx=4, pady=5
        self.frame_center_Check_drug_information.grid_propagate(False)

        #按钮设置

        # self.button1 = ttk.Button(self.window_Check_drug_information, text='编号查询', width=20, command=self.drug_id)
        # self.button1.place(x=50, y=485)

        self.button2 = ttk.Button(self.window_Check_drug_information, text='名称查询', width=20, command=self.drug_name)
        self.button2.place(x=225, y=350)
        #药品选择返回
        self.button3 = ttk.Button(self.window_Check_drug_information, text='返回', width=20,
                                  command=self.drug_selection_return)
        self.button3.place(x=600, y=350)
        
        # self.button4 = ttk.Button(self.window_Check_drug_information, text='药品信息数据报表', width=20,
        #                           command=self.drug_information_data_report)
        # self.button4.place(x=800, y=170)


        self.window_Check_drug_information.mainloop()  # 主消息循环
    
    
    def drug_selection_return(self):
        self.window_Check_drug_information.destroy()
    
    def drug_id(self):
        self.window_drug_id = tk.Tk()  # 初始框的声明
        self.window_drug_id.title('药品编号查询')
        self.window_drug_id.geometry('400x150')  # 1250x800

        tk.Label(self.window_drug_id, text='输入编号').place(x=80, y=40)
        self.var_usr_name = tk.StringVar()
        self.entry_usr_name = tk.Entry(self.window_drug_id, textvariable=self.var_usr_name)
        self.entry_usr_name.place(x=140, y=40)

        # 按钮设置
        btn_login = tk.Button(self.window_drug_id, text='查询', command=self.select_drug_id)
        btn_login.place(x=100, y=100)
        btn_sign_up = tk.Button(self.window_drug_id, text='返回', command=self.back_drugid)
        btn_sign_up.place(x=250, y=100)

        self.window_drug_id.mainloop()
        
    def select_drug_id(self):
        flag = True

        if str(self.entry_usr_name.get()) in self.id_Check_drug_information:

            db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root',
                                 password='123456')  # 打开数据库连接
            cursor = db.cursor()  # 使用cursor()方法获取操作游标
            sql = "SELECT * FROM medicine WHERE mno = '%s'" % (self.entry_usr_name.get())  # SQL 查询语句
            try:
                # 执行SQL语句
                cursor.execute(sql)
                # 获取所有记录列表
                results = cursor.fetchall()
                for row in results:
                    self.select_id_id_Check_drug_information.append(row[0])
                    self.select_id_name_Check_drug_information.append(row[1])
                    self.select_id_mode_Check_drug_information.append(row[2])
                    self.select_id_mefficacy_Check_drug_information.append(row[3])
                    self.select_id_mid_Check_drug_information.append(row[4])
                    # 打印结果
            except:
                print("Error")
                flag = False
                messagebox.showinfo('警告！', '错误！')
            db.close()  # 关闭数据库连接
            if flag:
                messagebox.showinfo('提示', '查询成功！')
                print(self.select_id_id_Check_drug_information)
                self.successful_drug_selection_query_interface(self.window_drug_id)

        else:
            messagebox.showinfo('警告！', '该药品编码不存在,请重新输入')
        
    def drug_name(self):
        self.window_drug_name= tk.Tk()  # 初始框的声明
        self.window_drug_name.title('药品名称查询')
        # self.window_drug_name.geometry('400x150')  # 1250x800
        
        self.frm1 = tk.Frame(self.window_drug_name)
        
        self.window_drug_name_mid=[]
        # tk.Label(self.window_drug_name, text='输入名称').place(x=80, y=40)
        self.var_usr_name = tk.StringVar()
        self.entry_usr_name = tk.Entry(self.frm1, textvariable=self.var_usr_name,width=35)
        
        self.drug_name_label = Label(self.frm1, text="药品名称:", font=('Verdana', 20))
        self.drug_name_label.grid(row=0, column=0, columnspan=2, sticky=NSEW, padx=50, pady=10) 
        # self.entry_usr_name.place(x=140, y=40)
        self.entry_usr_name.grid(column=0,row=1,padx=10)
        self.listbox1 = tk.Listbox(self.frm1)
        self.listbox1.grid(column=0, row=2, padx=10)
        
        self.listbox1.grid_remove()
        
        self.frm1.pack()
        self.entry_usr_name.bind("<KeyRelease>", self.entry_change)
        self.entry_usr_name.bind("<Button-3>", self.hide_list)
        
        self.listbox1.bind("<<ListboxSelect>>",self.list_select)
        
        self.window_drug_name_button1 = ttk.Button(self.frm1, text='确认', width=20,
                                            command=self.select_inventory)
        self.window_drug_name_button1.grid(row=3,column=0,padx=20, pady=10)

        
        self.window_drug_name.mainloop()
        
        
    

    
    def select_inventory(self):
        self.window_drug_name.destroy()
        self.window_select_inventory = tk.Tk()  # 初始框的声明
        self.window_select_inventory.geometry('400x400')
        self.window_select_inventory.title('选择库存')
        self.window_select_inventory_mid=[]
        self.var_select_inventory_input=tk.StringVar()
        
        db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root',
                                password='123456')  # 打开数据库连接
        cursor = db.cursor()  # 使用cursor()方法获取操作游标
        sql = "SELECT * FROM medicine WHERE mname = '%s'" % (self.list_drug_information[0])  # SQL 查询语句
        try:
            # 执行SQL语句
            cursor.execute(sql)
            # 获取所有记录列表
            results = cursor.fetchall()
            for row in results:
                self.window_select_inventory_mid.append(row[4])
                # 打印结果
        except:
            print("Error")
            flag = False
            messagebox.showinfo('警告！', '错误！')
        db.close()  # 关闭数据库连接
        
        
        
        

        self.window_select_inventory_label_text = tk.StringVar()
       
        self.window_select_inventory_label_text.set(', '.join(self.list_drug_information))
        print(self.list_drug_information)
        #self.window_select_inventory_label = tk.Label(self.window_select_inventory, textvariable=self.window_select_inventory_label_text, font=('Verdana', 20))
        self.window_select_inventory_text1=tk.Label(self.window_select_inventory,text='已选药品:' , font=('Verdana', 20))
        self.window_select_inventory_text1.grid(row=0,column=0)
        
        self.window_select_inventory_label1 = tk.Label(self.window_select_inventory,text=self.list_drug_information[0] , font=('Verdana', 20))
        self.window_select_inventory_label1.grid(row=0,column=1)
        
        self.window_select_inventory_text2=tk.Label(self.window_select_inventory,text='剩余库存:' , font=('Verdana', 20))
        self.window_select_inventory_text2.grid(row=1,column=0)
        
        self.window_select_inventory_label1 = tk.Label(self.window_select_inventory,text=self.window_select_inventory_mid , font=('Verdana', 20))
        self.window_select_inventory_label1.grid(row=1,column=1)
        
        self.window_select_inventory_label2 = tk.Label(self.window_select_inventory,text='选择个数:' , font=('Verdana', 20))
        self.window_select_inventory_label2.grid(row=2,column=0)
        
        self.window_select_inventory_input=tk.Entry(self.window_select_inventory,textvariable=self.var_select_inventory_input)
        self.window_select_inventory_input.grid(row=2,column=1)
        
        btn_login = tk.Button(self.window_select_inventory, text='确认', command=self.window_select_inventory_confirm)
        btn_login.grid(row=3,column=1)
        

        self.window_select_inventory.mainloop()
    
    
    
    
    

    def window_select_inventory_confirm(self):
        number=self.window_select_inventory_input.get()
        number=int(number)
        flag=True
        if number > int(self.window_select_inventory_mid[0]) or number <=0:
            messagebox.showinfo('警告！', '库存输入错误！')
            self.list_drug_information.clear()
            self.window_select_inventory.destory()
            
        else:
            remaining_inventory=int(self.window_select_inventory_mid[0])-number
                # 打开数据库连接
            db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root',
                                password='123456')
            cursor = db.cursor()  # 使用cursor()方法获取操作游标
            sql_update = "UPDATE medicine SET mid= '%s'\
                            WHERE mname = '%s'"  %(
                remaining_inventory, self.list_drug_information[0])  # SQL 插入语句
            try:
                cursor.execute(sql_update)  # 执行sql语句
                db.commit()  # 提交到数据库执行
            except:
                self.list_drug_information.clear()
                flag=False
                db.rollback()  # 发生错误时回滚
                messagebox.showerror('警告！', '更新失败，数据库连接失败！')
                db.close()  # 关闭数据库连接
        if flag:
            messagebox.showinfo('提示', f'{self.list_drug_information[0]}更新数据成功！剩余库存为{remaining_inventory}')
            self.select_drug_display.set(self.list_drug_information[0]+'x'+f'{number}')
            self.list_drug_information.clear()
            
              
        
        
        
    
    
    def entry_change(self, event):
        var_usr_name = self.entry_usr_name.get().strip()
        
        self.listbox1.delete(0, self.listbox1.size()-1)
        
        new_select_list = []
        for selection_info in self.name_Check_drug_information:
            if (len(var_usr_name) > 0) and (var_usr_name in selection_info):
                new_select_list.append(selection_info)
                
        print(var_usr_name, new_select_list)
        
        if len(new_select_list) > 0:
            self.set_list_value(new_select_list)
            # 显示出来
            self.listbox1.grid()
        else:
            self.listbox1.grid_remove()
    
         
    def hide_list(self, event):
        self.listbox1.grid_remove()
     
    #选择完下拉列表后，直接隐藏
    def list_select(self, event):
        index_num = self.listbox1.curselection()[0]
        select_value = self.listbox1.get(index_num)
        self.entry_usr_name.delete(0, tk.END)
        self.entry_usr_name.insert(tk.END, select_value)
        self.list_drug_information.append(self.entry_usr_name.get())
        # 隐藏下拉列表框
        
        self.listbox1.grid_remove()

         
    def set_list_value(self, value_list):
        #["A","AB","ABCD","ABCDE","ABCDEF"]
        for item in value_list:
            self.listbox1.insert("end",item)
        
        
        
        

        # 按钮设置
        # btn_login = tk.Button(self.window_drug_name, text='查询', command=self.select_drug_name)
        # btn_login.place(x=100, y=100)
        # btn_sign_up = tk.Button(self.window_drug_name, text='返回', command=self.back_drugname)
        # btn_sign_up.place(x=250, y=100)
        
        
        
    def back_drugname(self):
        self.window_drug_name.destroy()  # 回到查询界面
    
    def back_drugid(self):
        self.window_drug_id.destroy()  # 回到查询界面
        
        
        
        
    def select_drug_name(self):
        flag = True

        if str(self.entry_usr_name.get()) in self.name_Check_drug_information:

            db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root',
                                 password='123456')  # 打开数据库连接
            cursor = db.cursor()  # 使用cursor()方法获取操作游标
            sql = "SELECT * FROM medicine WHERE mname = '%s'" % (self.entry_usr_name.get())  # SQL 查询语句
            try:
                # 执行SQL语句
                cursor.execute(sql)
                # 获取所有记录列表
                results = cursor.fetchall()
                for row in results:
                    self.select_id_id_Check_drug_information.append(row[0])
                    self.select_id_name_Check_drug_information.append(row[1])
                    self.select_id_mode_Check_drug_information.append(row[2])
                    self.select_id_mefficacy_Check_drug_information.append(row[3])
                    self.select_id_mid_Check_drug_information.append(row[4])
                    # 打印结果
            except:
                print("Error")
                flag = False
                messagebox.showinfo('警告！', '错误！')
            db.close()  # 关闭数据库连接
            if flag:
                messagebox.showinfo('提示', '查询成功！')
                print(self.select_id_id_Check_drug_information)
                self.successful_drug_selection_query_interface(self.window_drug_name)

        else:
            messagebox.showinfo('警告！', '该药品名称不存在,请重新输入')
            
            
            
            
            
            
            
    #选择药品信息界面
    def successful_drug_selection_query_interface(self,parent_window):
        parent_window.destroy()
        self.select_drug_information_interface = tk.Tk()
        self.select_drug_information_interface.title('查询记录')
        self.select_drug_information_interface.geometry('900x360')

        self.frame_center_id = tk.Frame(self.select_drug_information_interface, width=1600, height=300)

        self.columns_id = (
            "编号", "名称", "服用方法", "功效")
        self.tree_id = ttk.Treeview(self.frame_center_id, show="headings", height=30, columns=self.columns_id)  # 18
        self.vbar_id = ttk.Scrollbar(self.frame_center_id, orient=VERTICAL, command=self.tree.yview)
        # 定义树形结构与滚动条
        self.tree_id.configure(yscrollcommand=self.vbar.set)

        # 表格的标题
        self.tree_id.column("编号", width=150, minwidth=100, anchor='center')  # 表示列,不显示
        self.tree_id.column("名称", width=150, minwidth=100, anchor='center')
        self.tree_id.column("服用方法", width=100, minwidth=100, anchor='center')
        self.tree_id.column("功效", width=450, minwidth=100, anchor='center')

        self.tree_id.heading('编号', text="编号", anchor='center')
        self.tree_id.heading("名称", text="名称", anchor='center')
        self.tree_id.heading("服用方法", text="服用方法", anchor='center')
        self.tree_id.heading("功效", text="功效", anchor='center')

        # 调用方法获取表格内容插入\
        self.tree_id.grid(row=0, column=0, sticky=NSEW)
        self.vbar_id.grid(row=0, column=1, sticky=NS)

        # 约束表格不重复插入
        # if str(self.entry_usr_name.get()) not in self.id_id:

        for i in range(min(len(self.select_id_id_Check_drug_information), len(self.select_id_name_Check_drug_information),
                           len(self.select_id_mode_Check_drug_information),
                           len(self.select_id_mefficacy_Check_drug_information))):  # 写入数据
            self.tree_id.insert('', i, values=(self.select_id_id_Check_drug_information[i], self.select_id_name_Check_drug_information[i],
                                            self.select_id_mode_Check_drug_information[i], self.select_id_mefficacy_Check_drug_information[i],
                                            ))


        for col in self.columns_id:  # 绑定函数，使表头可排序
            self.tree_id.heading(col, text=col,
                                 command=lambda _col=col: self.tree_sort_column(self.tree, _col, False))


        self.frame_center_id.grid(row=1, column=0, columnspan=2, padx=4, pady=5)
        self.frame_center_id.grid_propagate(False)


        self.window_id.mainloop()
        self.frame_center_id.tkraise()  # 开始显示主菜单
        
        
    
    

#录入药品信息
class Drug_information_entry:
    def __init__(self, parent_window):
        self.window_1 = None
        self.total_ano = None
        self.row_info = None
        self.row = None
        self.col = None
        parent_window.destroy()  # 销毁主界面

        self.window = Tk()  # 初始框的声明
        self.window.title('药品信息')

        self.frame_left_top = tk.Frame(width=700, height=250)#700,200
        self.frame_right_top = tk.Frame(width=400, height=300)
        self.frame_center = tk.Frame(width=1000, height=300)
        self.frame_bottom = tk.Frame(width=950, height=250)

        # 定义下方中心列表区域
        self.columns = (
        "编号", "名称", "服用方法", "功效",'库存')
        self.tree = ttk.Treeview(self.frame_center, show="headings", height=10, columns=self.columns)  # 18
        self.vbar = ttk.Scrollbar(self.frame_center, orient=VERTICAL, command=self.tree.yview)
        # 定义树形结构与滚动条
        self.tree.configure(yscrollcommand=self.vbar.set)


        self.tree.column("编号", width=150,minwidth=100, anchor='center')  # 表示列,不显示
        self.tree.column("名称", width=150, minwidth=100,anchor='center')
        self.tree.column("服用方法", width=100, minwidth=100,anchor='center')
        self.tree.column("功效", width=450,minwidth=100, anchor='center')
        self.tree.column("库存", width=80,minwidth=100, anchor='center')

        self.tree.heading('编号', text="编号", anchor='center')
        self.tree.heading("名称", text="名称", anchor='center')
        self.tree.heading("服用方法", text="服用方法", anchor='center')
        self.tree.heading("功效", text="功效", anchor='center')
        self.tree.heading("库存", text="库存", anchor='center')

        self.tree.grid(row=0, column=0, sticky=NSEW)
        self.vbar.grid(row=0, column=1, sticky=NS)

        self.id = []
        self.name = []
        self.mode=[]
        self.mefficacy=[]
        self.mid=[]

        db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root',
                             password='123456')
        cursor = db.cursor()  # 使用cursor()方法获取操作游标
        sql = "SELECT * FROM medicine"  # SQL 查询语句 customer_data_copy1 customer_d


        try:
            # 执行SQL语句
            cursor.execute(sql)
            # 获取所有记录列表
            results = cursor.fetchall()
            for row in results:
                self.id.append(row[0])
                self.name.append(row[1])
                self.mode.append(row[2])
                self.mefficacy.append(row[3])
                self.mid.append(row[4])


        except:
            print("Error: unable to fetch data")
            messagebox.showinfo('警告！', '数据库连接失败！')
        db.close()  # 关闭数据库连接


        for i in range(min(len(self.id), len(self.name), len(self.mode), len(self.mefficacy),len(self.mid))):  # 写入数据
            self.tree.insert('', i, values=(self.id[i], self.name[i], self.mode[i], self.mefficacy[i],self.mid[i],
                                            ))
        for col in self.columns:  # 绑定函数，使表头可排序
            self.tree.heading(col, text=col,
                              command=lambda _col=col: self.tree_sort_column(self.tree, _col, False))

        # 编号
        self.top_title = Label(self.frame_left_top, text="药品信息:", font=('Verdana', 20))
        self.top_title.grid(row=0, column=0, columnspan=2, sticky=NSEW, padx=50, pady=10)  # 50 ,10

        # 定义左上方区域
        self.left_top_frame = tk.Frame(self.frame_left_top)

        
        self.var_id = StringVar()  # 声明编号
        self.var_name = StringVar()  # 声明名称
        self.var_mode = StringVar()  # 声明服用方法
        self.var_mefficacy = StringVar()  # 声明功效
        self.var_mid=StringVar()  # 声明库存
        # 编号
        self.right_top_id_label = Label(self.frame_left_top, text="编号：", font=('Verdana', 15))
        self.right_top_id_entry = Entry(self.frame_left_top, textvariable=self.var_id, font=('Verdana', 15))
        self.right_top_id_label.grid(row=1, column=0)  # 位置设置
        self.right_top_id_entry.grid(row=1, column=1)
        # 名称
        self.right_top_name_label = Label(self.frame_left_top, text="名称：", font=('Verdana', 15))
        self.right_top_name_entry = Entry(self.frame_left_top, textvariable=self.var_name, font=('Verdana', 15))
        self.right_top_name_label.grid(row=2, column=0)  # 位置设置
        self.right_top_name_entry.grid(row=2, column=1)
        # 服用方法
        self.right_top_gender_label = Label(self.frame_left_top, text="服用方法：", font=('Verdana', 15))
        self.right_top_gender_entry = Entry(self.frame_left_top, textvariable=self.var_mode,
                                            font=('Verdana', 15))
        self.right_top_gender_label.grid(row=3, column=0)  # 位置设置
        self.right_top_gender_entry.grid(row=3, column=1)
        # 功效
        self.right_top_gender_label = Label(self.frame_left_top, text="功效：", font=('Verdana', 15))
        self.right_top_gender_entry = Entry(self.frame_left_top, textvariable=self.var_mefficacy,
                                            font=('Verdana', 15))
        self.right_top_gender_label.grid(row=4, column=0)  # 位置设置
        self.right_top_gender_entry.grid(row=4, column=1)
        
        # 库存
        self.right_top_gender_label = Label(self.frame_left_top, text="库存：", font=('Verdana', 15))
        self.right_top_gender_entry = Entry(self.frame_left_top, textvariable=self.var_mid,
                                            font=('Verdana', 15))
        self.right_top_gender_label.grid(row=5, column=0)  # 位置设置
        self.right_top_gender_entry.grid(row=5, column=1)

        # 定义右上方区域
        self.right_top_title = Label(self.frame_right_top, text="操作：", font=('Verdana', 20))

        self.tree.bind('<Button-1>', self.click)  # 左键获取位置
        self.right_top_button1 = ttk.Button(self.frame_right_top, text='新增药品信息', width=20, command=self.new_row)
        self.right_top_button2 = ttk.Button(self.frame_right_top, text='更新药品信息', width=20,
                                            command=self.updata_row)
        self.right_top_button3 = ttk.Button(self.frame_right_top, text='删除选中药品信息', width=20,
                                            command=self.del_row)
        self.right_top_button4 = ttk.Button(self.frame_right_top, text='返回', width=20,
                                            command=self.back)

        self.right_top_title.grid(row=1, column=0, pady=10)
        self.right_top_button1.grid(row=2, column=0, padx=20, pady=10)
        self.right_top_button2.grid(row=3, column=0, padx=20, pady=10)
        self.right_top_button3.grid(row=4, column=0, padx=20, pady=10)
        self.right_top_button4.grid(row=5, column=0, padx=20, pady=10)



        self.frame_left_top.grid(row=0, column=0, padx=2, pady=5)
        self.frame_right_top.grid(row=0, column=1, padx=30, pady=30)
        self.frame_center.grid(row=1, column=0, columnspan=3, padx=0, pady=0)  #, padx=4, pady=5
        # self.frame_bottom.grid(row=2, column=0, columnspan=2)


        self.frame_left_top.grid_propagate(False)
        self.frame_right_top.grid_propagate(False)
        self.frame_center.grid_propagate(False)
        # self.frame_bottom.grid_propagate(False)

        self.frame_left_top.tkraise()  # 开始显示主菜单
        self.frame_right_top.tkraise()  # 开始显示主菜单
        self.frame_center.tkraise()  # 开始显示主菜单
        # self.frame_bottom.tkraise()  # 开始显示主菜单

        self.window.protocol("WM_DELETE_WINDOW", self.back)  # 捕捉右上角关闭点击，回到首页


    def tree_sort_column(self, tv, col, reverse):  # Treeview、列名、排列方式
        l = [(tv.set(k, col), k) for k in tv.get_children('')]
        l.sort(reverse=reverse)  # 排序方式
        # rearrange items in sorted positions
        for index, (val, k) in enumerate(l):  # 根据排序后索引移动
            tv.move(k, '', index)
        tv.heading(col, command=lambda: self.tree_sort_column(tv, col, not reverse))  # 重写标题，使之成为再点倒序的标题
    
    def back(self):
        AgencyManage(self.window)  #回退至经办人操作窗口

    def customer_1(self):
        pass

    def click(self,event):#click点击 将列表的数据显示在输入框内
        self.col = self.tree.identify_column(event.x)  # 列
        self.row = self.tree.identify_row(event.y)  # 行

        print(self.col)
        print(self.row)
        self.row_info = self.tree.item(self.row, "values")
        self.var_id.set(self.row_info[0])
        self.var_name.set(self.row_info[1])
        self.var_mode.set(self.row_info[2])
        self.var_mefficacy.set(self.row_info[3])
        self.var_mid.set(self.row_info[4])
        self.right_top_id_entry = Entry(self.frame_left_top, state='disabled', textvariable=self.var_id,
                                        font=('Verdana', 15))
        print()

    def new_row(self):
        flag = True
        flag_1=True
        flag_please_fill_in_customer_data=True
        # if str(self.var_id.get()) in self.id:
        print()
        
        if str(self.var_id.get()) in self.id:
            messagebox.showinfo('警告！', '该药品编号信息已存在！')
        else:
                # 打开数据库连接
            if int(self.var_mid.get()) <0 :
                messagebox.showinfo('警告！', '库存信息有误！')
                flag_1=FALSE
                flag_please_fill_in_customer_data=False
                flag=False
                    
            if flag_1:
                if flag:
                    db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root',
                                        password='123456')
                    cursor = db.cursor()  # 使用cursor()方法获取操作游标

                    sql = "INSERT INTO medicine(mno, mname, mmode, mefficacy,mid) \
                                    VALUES ('%s', '%s', '%s', '%s','%s')" % \
                        (self.var_id.get(), self.var_name.get(), self.var_mode.get(), self.var_mefficacy.get(),self.var_mid.get())  # SQL 插入语句

                    try:
                        cursor.execute(sql)  # 执行sql语句
                        db.commit()  # 提交到数据库执行
                    except:
                        flag = False
                        db.rollback()  # 发生错误时回滚
                        messagebox.showinfo('警告！', '数据库连接失败！')
                    db.close()  # 关闭数据库连接


            if flag:
                self.id.append(self.var_id.get())
                self.name.append(self.var_name.get())
                self.mode.append(self.var_mode.get())
                self.mefficacy.append(self.var_mefficacy.get())
                self.mid.append(self.var_mid.get())
                self.tree.insert('', len(self.id) - 1, values=(
                    self.id[len(self.id) - 1], self.name[len(self.id) - 1], self.mode[len(self.id) - 1],
                    self.mefficacy[len(self.id) - 1],
                    self.mid[len(self.mid) - 1],
                )),
                self.tree.update()
                messagebox.showinfo('提示！', '新建药品信息成功！')
         




    def updata_row(self):
        flag=True
        res = messagebox.askokcancel('警告！', '是否更新所填数据？')#按钮返回true或flase
        if res:
            if self.var_id.get() == self.row_info[0]:  # 如果所填学号 与 所选学号一致
                # 打开数据库连接
                db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root',
                                     password='123456')
                cursor = db.cursor()  # 使用cursor()方法获取操作游标
                sql_update = "UPDATE medicine SET mno = '%s', mname = '%s', mmode = '%s' ,mefficacy= '%s',mid= '%s'\
                                 WHERE mno = '%s'"  %(
                    self.var_id.get(), self.var_name.get(), self.var_mode.get(), self.var_mefficacy.get(),self.var_mid.get(),self.var_id.get())  # SQL 插入语句
                try:
                    cursor.execute(sql_update)  # 执行sql语句
                    db.commit()  # 提交到数据库执行
                    messagebox.showinfo('提示！', '更新成功！')
                except:
                    flag=False
                    db.rollback()  # 发生错误时回滚
                    messagebox.showerror('警告！', '更新失败，数据库连接失败！')
                    db.close()  # 关闭数据库连接
                if flag:
                    id_index = self.id.index(self.row_info[0])
                    self.name[id_index] = self.var_name.get()
                    self.id[id_index] = self.var_id.get()
                    self.mode[id_index] = self.var_mode.get()
                    self.mefficacy[id_index]=self.var_mefficacy.get()
                    self.mid[id_index]=self.var_mid.get()
                    

                    self.tree.item(self.tree.selection()[0], values=(
                        self.var_id.get(), self.var_name.get(), self.var_mode.get()
                        ,self.var_mefficacy.get(),self.var_mid.get()))  # 修改对于行信息
                else:
                    messagebox.showinfo('警告！', '该药品编号不存在')
            else:
                messagebox.showinfo('警告！', '不能修改药品编号！')


    def del_row(self):
        res = messagebox.askokcancel('警告！', '是否删除所选数据？')
        if res:
            # 打开数据库连接
            db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root',
                                 password='123456')
            cursor = db.cursor()  # 使用cursor()方法获取操作游标
            sql_delete = "DELETE FROM medicine WHERE mno = '%s'" % (self.row_info[0])  # SQL 插入语句
            try:
                cursor.execute(sql_delete)  # 执行sql语句
                db.commit()  # 提交到数据库执行
                messagebox.showinfo('提示！', '删除成功！')
            except:
                db.rollback()  # 发生错误时回滚
                messagebox.showinfo('警告！', '删除失败，数据库连接失败！')
            db.close()  # 关闭数据库连接

            id_index = self.id.index(self.row_info[0])

            del self.id[id_index]
            del self.name[id_index]
            del self.mode[id_index]
            del self.mefficacy[id_index]
            del self.mid[id_index]

            self.tree.delete(self.tree.selection()[0])  # 删除所选行


#查看顾客信息

class Customer_select:
    def __init__(self,parent_window):
        #定义以id方式查询的字段
        self.vbar_name = None
        self.tree_name = None
        self.columns_name = None
        self.frame_center_name = None
        self.window_name = None
        self.columns_id = None
        self.vbar_id = None
        self.tree_id = None
        self.frame_center_id = None
        self.window_id = None

        self.id_id = []
        self.name_id = []
        self.gender_id = []
        self.age_id = []
        self.address_id = []
        self.phone_id = []
        self.symptom_id = []
        self.mno_id = []
        self.ano_id = []
        self.date_id = []
        self.remark_id = []

        #备份字段:
        self.backup_id=[]
        self.backup_name=[]

        self.entry_usr_name = None
        self.var_usr_name = None
        parent_window.destroy()  # 销毁子界面
        self.window_test = tk.Tk()  # 初始框的声明
        self.window_test.title('查看患者信息')
        self.window_test.geometry('1250x600')#1250x800
        # frame1=tk.Frame(self.window)
        # 画布功能添加图片 目前仅支持gif格式的图片

        canvas = tk.Canvas(self.window_test, height=200, width=750)
        imge_file = tk.PhotoImage(file='test_数据库\顾客查询.gif')
        imge = canvas.create_image(0, 0, anchor='nw', image=imge_file)
        canvas.grid(row=0, column=0,padx=120)
        # lamada 相当于定义函数/
        self.frame_center = tk.Frame(width=1600, height=300)

        self.columns = (
        "编号", "姓名", "性别", "年龄", "住址", "电话", "症状", "已购药品", "经办人", "录入日期", "备注")
        self.tree = ttk.Treeview(self.frame_center, show="headings", height=30, columns=self.columns)  # 18
        self.vbar = ttk.Scrollbar(self.frame_center, orient=VERTICAL, command=self.tree.yview)
        # 定义树形结构与滚动条
        self.tree.configure(yscrollcommand=self.vbar.set)

        # 表格的标题
        self.tree.column("编号", width=150, anchor='center')  # 表示列,不显示
        self.tree.column("姓名", width=150, anchor='center')
        self.tree.column("性别", width=100, anchor='center')
        self.tree.column("年龄", width=100, anchor='center')
        self.tree.column("住址", width=100, anchor='center')
        self.tree.column("电话", width=100, anchor='center')
        self.tree.column("症状", width=100, anchor='center')
        self.tree.column("已购药品", width=100, anchor='center')
        self.tree.column("经办人", width=100, anchor='center')
        self.tree.column("录入日期", width=100, anchor='center')
        self.tree.column("备注", width=100, anchor='center')

        # 调用方法获取表格内容插入
        self.tree.grid(row=0, column=0, sticky=NSEW)
        self.vbar.grid(row=0, column=1, sticky=NS)

        self.id = []
        self.name = []
        self.gender = []
        self.age = []
        self.address = []
        self.phone = []
        self.symptom = []
        self.mno = []
        self.ano = []
        self.date = []
        self.remark = []
        # 打开数据库连接
        db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root',
                             password='123456')
        cursor = db.cursor()  # 使用cursor()方法获取操作游标
        sql = "SELECT * FROM customer_data_copy1"  # SQL 查询语句 customer_data_copy1 customer_d
        try:
            # 执行SQL语句
            cursor.execute(sql)
            # 获取所有记录列表
            results = cursor.fetchall()
            for row in results:
                self.id.append(row[0])
                self.name.append(row[1])
                self.gender.append(row[2])
                self.age.append(row[3])
                self.address.append(row[4])
                self.phone.append(row[5])
                self.symptom.append(row[6])
                self.mno.append(row[7])
                self.ano.append(row[8])
                self.date.append(row[9])
                self.remark.append(row[10])


        except:
            print("Error: unable to fetch data")
            messagebox.showinfo('警告！', '数据库连接失败！')
        db.close()  # 关闭数据库连接

        print("test***********************")
        for i in range(min(len(self.id), len(self.name), len(self.gender), len(self.phone))):  # 写入数据
            self.tree.insert('', i, values=(
            self.id[i], self.name[i], self.gender[i], self.age[i], self.address[i], self.phone[i], self.symptom[i],
            self.mno[i], self.ano[i], self.date[i], self.remark[i]))
        for col in self.columns:  # 绑定函数，使表头可排序
            self.tree.heading(col, text=col,
                              command=lambda _col=col: self.tree_sort_column(self.tree, _col, False))



        self.frame_center.grid_propagate(False)
        self.frame_center.grid(row=2, column=0, columnspan=2, padx=4, pady=5)
        #按钮设置
        self.button1 = ttk.Button(self.window_test, text='编号查询', width=20, command=self.Customer_id)
        self.button1.place(x=175,y=525)

        self.button2 = ttk.Button(self.window_test, text='姓名查询', width=20, command=self.Customer_name)
        self.button2.place(x=575,y=525)

        self.button3 = ttk.Button(self.window_test, text='返回', width=20, command=lambda:AgencyManage(self.window_test))
        self.button3.place(x=975, y=525)
        
        self.button4 = ttk.Button(self.window_test, text='患者信息数据报表', width=20, command=self.customer_information_data_report)
        self.button4.place(x=1000, y=170)

        self.frame_center.tkraise()  # 开始显示主菜单
        self.window_test.mainloop()
        
    def customer_information_data_report(self):
        host = "localhost"
        username = "root"
        passwd = "123456"
        db = "medicine management system"
        port = 3306
        charset = "utf8"
        excel_name = "患者信息.xlsx"
        table_name='customer_data_copy1'
        sheet=[]
        conn = pymysql.connect(host=host, user=username,
                        passwd=passwd, db=db,
                        port=port, charset=charset)
        cur = conn.cursor()
        sql = "select * from %s;" % table_name
        cur.execute(sql)
        sql_result = cur.fetchall()
        cur.close()
        conn.close()
        # 写 Excel
        book = openpyxl.Workbook()
        sheet = book.active
        # fff = [filed[0] for filed in cur.description]  # 获取表头信息
        fff=['编号','姓名','性别','年龄','住址','电话','症状','已购药品','经办人','录入日期','备注']
        sheet.append(fff) # type: ignore
        
        for i in sql_result:
            sheet.append(i) # type: ignore
        
        try:
            book.save("%s" % excel_name)
        except:
            messagebox.showinfo('警告！', '数据报表失败!')
            
        path_=os.getcwd()
        messagebox.showinfo('提示！', f'患者信息数据报表创建成功!保存路径为:{path_}')

    def tree_sort_column(self, tree, _col, param):
        pass

    def back(self):
        AgencyManage(self.window)  # 回退至经办人操作窗口

    def back_select_id(self):
        self.window_id.destroy()
    def back_id(self):
        self.window.destroy()#回到查询界面

            #按编号查询顾客信息
    def Customer_id(self):
        self.window = tk.Tk()  # 初始框的声明
        self.window.title('患者编号查询')
        self.window.geometry('400x150')  # 1250x800

        tk.Label(self.window, text='输入编号').place(x=80, y=40)
        self.var_usr_name = tk.StringVar()
        self.entry_usr_name = tk.Entry(self.window, textvariable=self.var_usr_name)
        self.entry_usr_name.place(x=140, y=40)

        #按钮设置
        btn_login = tk.Button(self.window, text='查询', command=self.select_id)
        btn_login.place(x=100, y=100)
        btn_sign_up = tk.Button(self.window, text='返回', command=self.back_id)
        btn_sign_up.place(x=250, y=100)


        self.window.mainloop()


    #按名字查询顾客信息
    def Customer_name(self):
        self.window = tk.Tk()  # 初始框的声明
        self.window.title('患者编号查询')
        self.window.geometry('400x150')  # 1250x800

        tk.Label(self.window, text='输入姓名').place(x=80, y=40)
        self.var_usr_name = tk.StringVar()
        self.entry_usr_name = tk.Entry(self.window, textvariable=self.var_usr_name)
        self.entry_usr_name.place(x=140, y=40)

        # 按钮设置
        btn_login = tk.Button(self.window, text='查询', command=self.select_name)
        btn_login.place(x=100, y=100)
        btn_sign_up = tk.Button(self.window, text='返回', command=self.back_id)
        btn_sign_up.place(x=250, y=100)


#姓名查询
    def select_name(self):
        flag = True

        if str(self.entry_usr_name.get()) in self.name:

            db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root',
                                 password='123456')  # 打开数据库连接
            cursor = db.cursor()  # 使用cursor()方法获取操作游标
            sql = "SELECT * FROM customer_data_copy1 WHERE cname = '%s'" % (self.entry_usr_name.get())  # SQL 查询语句
            try:
                # 执行SQL语句
                cursor.execute(sql)
                # 获取所有记录列表
                results = cursor.fetchall()
                for row in results:
                    self.id_id.append(row[0])
                    self.name_id.append(row[1])
                    self.gender_id.append(row[2])
                    self.age_id.append(row[3])
                    self.address_id.append(row[4])
                    self.phone_id.append(row[5])
                    self.symptom_id.append(row[6])
                    self.mno_id.append(row[7])
                    self.ano_id.append(row[8])
                    self.date_id.append(row[9])
                    self.remark_id.append(row[10])
                    # 打印结果
            except:
                print("Error")
                flag = False
                messagebox.showinfo('警告！', '错误！')
            db.close()  # 关闭数据库连接
            if flag:
                messagebox.showinfo('提示', '查询成功！')
                self.select_name_gui(self.window)

        else:
            messagebox.showinfo('警告！', '该患者姓名不存在,请重新输入')

    #id查询方式实现
    def select_id(self):
        flag=True

        if str(self.entry_usr_name.get()) in self.id:

            db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root',
                                 password='123456')  # 打开数据库连接
            cursor = db.cursor()  # 使用cursor()方法获取操作游标
            sql = "SELECT * FROM customer_data_copy1 WHERE cno = '%s'" % (self.entry_usr_name.get())  # SQL 查询语句
            try:
                # 执行SQL语句
                cursor.execute(sql)
                # 获取所有记录列表
                results = cursor.fetchall()
                for row in results:
                    self.backup_id.append(row[0])
                    self.id_id.append(row[0])
                    self.name_id.append(row[1])
                    self.gender_id.append(row[2])
                    self.age_id.append(row[3])
                    self.address_id.append(row[4])
                    self.phone_id.append(row[5])
                    self.symptom_id.append(row[6])
                    self.mno_id.append(row[7])
                    self.ano_id.append(row[8])
                    self.date_id.append(row[9])
                    self.remark_id.append(row[10])
                    # 打印结果
            except:
                print("Error")
                flag=False
                messagebox.showinfo('警告！', '错误！')
            db.close()  # 关闭数据库连接
            if flag:
                messagebox.showinfo('提示', '查询成功！')
                self.select_id_gui(self.window)

        else:
            messagebox.showinfo('警告！', '该患者编码不存在,请重新输入')




    # 查询id gui界面
    def select_id_gui(self,parent_window):
        parent_window.destroy()  # 销毁主界面
        self.window_id = tk.Tk()
        self.window_id.title('查询记录')
        self.window_id.geometry('1250x400')

        self.frame_center_id = tk.Frame(self.window_id,width=1600, height=300)

        self.columns_id = (
            "编号", "姓名", "性别", "年龄", "住址", "电话", "症状", "已购药品", "经办人", "录入日期", "备注")
        self.tree_id = ttk.Treeview(self.frame_center_id, show="headings", height=30, columns=self.columns_id)  # 18
        self.vbar_id = ttk.Scrollbar(self.frame_center_id, orient=VERTICAL, command=self.tree.yview)
        # 定义树形结构与滚动条
        self.tree_id.configure(yscrollcommand=self.vbar.set)

        # 表格的标题
        self.tree_id.column("编号", width=150, anchor='center')  # 表示列,不显示
        self.tree_id.column("姓名", width=150, anchor='center')
        self.tree_id.column("性别", width=100, anchor='center')
        self.tree_id.column("年龄", width=100, anchor='center')
        self.tree_id.column("住址", width=100, anchor='center')
        self.tree_id.column("电话", width=100, anchor='center')
        self.tree_id.column("症状", width=100, anchor='center')
        self.tree_id.column("已购药品", width=100, anchor='center')
        self.tree_id.column("经办人", width=100, anchor='center')
        self.tree_id.column("录入日期", width=100, anchor='center')
        self.tree_id.column("备注", width=100, anchor='center')

        # 调用方法获取表格内容插入
        self.tree_id.grid(row=0, column=0, sticky=NSEW)
        self.vbar_id.grid(row=0, column=1, sticky=NS)

        #约束表格不重复插入
        # if str(self.entry_usr_name.get()) not in self.id_id:

        for i in range(min(len(self.id_id), len(self.name_id), len(self.gender_id), len(self.phone_id))):  # 写入数据
            self.tree_id.insert('', i, values=(
            self.id_id[i], self.name_id[i], self.gender_id[i], self.age_id[i], self.address_id[i], self.phone_id[i], self.symptom_id[i],
            self.mno_id[i], self.ano_id[i], self.date_id[i], self.remark_id[i]))
        for col in self.columns:  # 绑定函数，使表头可排序
            self.tree_id.heading(col, text=col,
                              command=lambda _col=col: self.tree_sort_column(self.tree, _col, False))

        self.frame_center_id.grid_propagate(False)
        self.frame_center_id.grid(row=1, column=0, columnspan=2, padx=4, pady=5)

        self.frame_center_id.tkraise()  # 开始显示主菜单
        self.window_id.mainloop()
        self.window_id.protocol("WM_DELETE_WINDOW", self.back_select_id)  # 捕捉右上角关闭点击，回到首页






    #查询姓名gui界面
    def select_name_gui(self,parent_window):
        parent_window.destroy()  # 销毁主界面
        self.window_name = tk.Tk()
        self.window_name.title('查询记录')
        self.window_name.geometry('1250x400')

        self.frame_center_name = tk.Frame(self.window_name, width=1600, height=300)

        self.columns_name = (
            "编号", "姓名", "性别", "年龄", "住址", "电话", "症状", "已购药品", "经办人", "录入日期", "备注")
        self.tree_name = ttk.Treeview(self.frame_center_name, show="headings", height=30, columns=self.columns_name)  # 18
        self.vbar_name = ttk.Scrollbar(self.frame_center_name, orient=VERTICAL, command=self.tree.yview)
        # 定义树形结构与滚动条
        self.tree_name.configure(yscrollcommand=self.vbar.set)

        # 表格的标题
        self.tree_name.column("编号", width=150, anchor='center')  # 表示列,不显示
        self.tree_name.column("姓名", width=150, anchor='center')
        self.tree_name.column("性别", width=100, anchor='center')
        self.tree_name.column("年龄", width=100, anchor='center')
        self.tree_name.column("住址", width=100, anchor='center')
        self.tree_name.column("电话", width=100, anchor='center')
        self.tree_name.column("症状", width=100, anchor='center')
        self.tree_name.column("已购药品", width=100, anchor='center')
        self.tree_name.column("经办人", width=100, anchor='center')
        self.tree_name.column("录入日期", width=100, anchor='center')
        self.tree_name.column("备注", width=100, anchor='center')

        # 调用方法获取表格内容插入
        self.tree_name.grid(row=0, column=0, sticky=NSEW)
        self.vbar_name.grid(row=0, column=1, sticky=NS)


        # 约束表格不重复插入
        for i in range(min(len(self.id_id), len(self.name_id), len(self.gender_id), len(self.phone_id))):  # 写入数据
            self.tree_name.insert('', i, values=(
                self.id_id[i], self.name_id[i], self.gender_id[i], self.age_id[i], self.address_id[i], self.phone_id[i],
                self.symptom_id[i],
                self.mno_id[i], self.ano_id[i], self.date_id[i], self.remark_id[i]))
        for col in self.columns:  # 绑定函数，使表头可排序
            self.tree_name.heading(col, text=col,
                                 command=lambda _col=col: self.tree_sort_column(self.tree, _col, False))

        self.frame_center_name.grid_propagate(False)
        self.frame_center_name.grid(row=1, column=0, columnspan=2, padx=4, pady=5)

        self.window_name.mainloop()
        self.frame_center_name.tkraise()  # 开始显示主菜单
        self.window_name.protocol("WM_DELETE_WINDOW", self.back)  # 捕捉右上角关闭点击，回到首页

#查看药品信息
class Check_drug_information:
    def __init__(self,parent_window):
        parent_window.destroy()  # 销毁子界面
        self.window_Check_drug_information = tk.Tk()  # 初始框的声明
        self.window_Check_drug_information.title('查看药品信息')
        self.window_Check_drug_information.geometry('1000x600')

        # 画布功能添加图片 目前仅支持gif格式的图片  test_数据库\222222.gif
        canvas = tk.Canvas(self.window_Check_drug_information, height=200, width=500)
        imge_file = tk.PhotoImage(file='test_数据库\药品信息_500x281.gif')
        imge = canvas.create_image(0, 0, anchor='nw', image=imge_file)
        canvas.grid(row=0, column=0,padx=250)
        # lamada 相当于定义函数

        self.frame_center_Check_drug_information = tk.Frame(self.window_Check_drug_information,width=1000, height=300)

        self.columns_Check_drug_information = (
            "编号", "名称", "服用方法", "功效",'库存')
        self.tree = ttk.Treeview(self.frame_center_Check_drug_information, show="headings", height=10, columns=self.columns_Check_drug_information)  # 18
        self.vbar = ttk.Scrollbar(self.frame_center_Check_drug_information, orient=VERTICAL, command=self.tree.yview)
        # 定义树形结构与滚动条
        self.tree.configure(yscrollcommand=self.vbar.set)

        self.tree.column("编号", width=150, minwidth=100, anchor='center')  # 表示列,不显示
        self.tree.column("名称", width=150, minwidth=100, anchor='center')
        self.tree.column("服用方法", width=100, minwidth=100, anchor='center')
        self.tree.column("功效", width=450, minwidth=100, anchor='center')
        self.tree.column("库存", width=80,minwidth=100, anchor='center')

        self.tree.heading('编号', text="编号", anchor='center')
        self.tree.heading("名称", text="名称", anchor='center')
        self.tree.heading("服用方法", text="服用方法", anchor='center')
        self.tree.heading("功效", text="功效", anchor='center')
        self.tree.heading("库存", text="库存", anchor='center')

        self.tree.grid(row=0, column=0, sticky=NSEW)
        self.vbar.grid(row=0, column=1, sticky=NS)

        #开始查询框的变量
        self.id_Check_drug_information = []
        self.name_Check_drug_information = []
        self.mode_Check_drug_information = []
        self.mefficacy_Check_drug_information = []
        self.mmid_Check_drug_information=[]

        #药品编号查询变量：
        self.select_id_mefficacy_Check_drug_information=[]
        self.select_id_mode_Check_drug_information=[]
        self.select_id_name_Check_drug_information=[]
        self.select_id_id_Check_drug_information=[]
        
        #数据库查询
        db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root',
                             password='123456')
        cursor = db.cursor()  # 使用cursor()方法获取操作游标
        sql = "SELECT * FROM medicine"  # SQL 查询语句 customer_data_copy1 customer_d
        try:
            # 执行SQL语句
            cursor.execute(sql)
            # 获取所有记录列表
            results = cursor.fetchall()
            for row in results:
                self.id_Check_drug_information.append(row[0])
                self.name_Check_drug_information.append(row[1])
                self.mode_Check_drug_information.append(row[2])
                self.mefficacy_Check_drug_information.append(row[3])
                self.mmid_Check_drug_information.append(row[4])


        except:
            print("Error: unable to fetch data")
            messagebox.showinfo('警告！', '数据库连接失败！')
        db.close()  # 关闭数据库连接

        for i in range(min(len(self.id_Check_drug_information), len(self.name_Check_drug_information), len(self.mode_Check_drug_information), len(self.mefficacy_Check_drug_information))):  # 写入数据
            self.tree.insert('', i, values=(self.id_Check_drug_information[i], self.name_Check_drug_information[i], self.mode_Check_drug_information[i], self.mefficacy_Check_drug_information[i],self.mmid_Check_drug_information[i],
                                            ))

        #药品信息框
        self.frame_center_Check_drug_information.grid(row=1, column=0, columnspan=3, padx=0, pady=0)  # , padx=4, pady=5
        self.frame_center_Check_drug_information.grid_propagate(False)

        #按钮设置

        self.button1 = ttk.Button(self.window_Check_drug_information, text='编号查询', width=20, command=self.drug_id)
        self.button1.place(x=50, y=485)

        self.button2 = ttk.Button(self.window_Check_drug_information, text='名称查询', width=20, command=self.drug_name)
        self.button2.place(x=425, y=485)

        self.button3 = ttk.Button(self.window_Check_drug_information, text='返回', width=20,
                                  command=lambda: AgencyManage(self.window_Check_drug_information))
        self.button3.place(x=800, y=485)
        
        self.button4 = ttk.Button(self.window_Check_drug_information, text='药品信息数据报表', width=20,
                                  command=self.drug_information_data_report)
        self.button4.place(x=800, y=170)


        self.window_Check_drug_information.mainloop()  # 主消息循环
        
        
    #数据报表
    def drug_information_data_report(self):
        host = "localhost"
        username = "root"
        passwd = "123456"
        db = "medicine management system"
        port = 3306
        charset = "utf8"
        excel_name = "药品信息.xlsx"
        table_name='medicine'
        sheet=[]
        conn = pymysql.connect(host=host, user=username,
                        passwd=passwd, db=db,
                        port=port, charset=charset)
        cur = conn.cursor()
        sql = "select * from %s;" % table_name
        cur.execute(sql)
        sql_result = cur.fetchall()
        cur.close()
        conn.close()
        # 写 Excel
        book = openpyxl.Workbook()
        sheet = book.active
        # fff = [filed[0] for filed in cur.description]  # 获取表头信息
        fff=['药品编号','药品名称','内服/外服','作用','库存']
        sheet.append(fff) # type: ignore
        
        for i in sql_result:
            sheet.append(i) # type: ignore
        
        try:
            book.save("%s" % excel_name)
        except:
            messagebox.showinfo('警告！', '数据报表失败!')
            
        path_=os.getcwd()
        messagebox.showinfo('提示！', f'药品信息数据报表创建成功!保存路径为:{path_}')
    
    def drug_id(self):
        self.window = tk.Tk()  # 初始框的声明
        self.window.title('药品编号查询')
        self.window.geometry('400x150')  # 1250x800

        tk.Label(self.window, text='输入编号').place(x=80, y=40)
        self.var_usr_name = tk.StringVar()
        self.entry_usr_name = tk.Entry(self.window, textvariable=self.var_usr_name)
        self.entry_usr_name.place(x=140, y=40)

        # 按钮设置
        btn_login = tk.Button(self.window, text='查询', command=self.select_drug_id)
        btn_login.place(x=100, y=100)
        btn_sign_up = tk.Button(self.window, text='返回', command=self.back_drug)
        btn_sign_up.place(x=250, y=100)

        self.window.mainloop()

    def select_drug_id(self):
        flag = True

        if str(self.entry_usr_name.get()) in self.id_Check_drug_information:

            db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root',
                                 password='123456')  # 打开数据库连接
            cursor = db.cursor()  # 使用cursor()方法获取操作游标
            sql = "SELECT * FROM medicine WHERE mno = '%s'" % (self.entry_usr_name.get())  # SQL 查询语句
            try:
                # 执行SQL语句
                cursor.execute(sql)
                # 获取所有记录列表
                results = cursor.fetchall()
                for row in results:
                    self.select_id_id_Check_drug_information.append(row[0])
                    self.select_id_name_Check_drug_information.append(row[1])
                    self.select_id_mode_Check_drug_information.append(row[2])
                    self.select_id_mefficacy_Check_drug_information.append(row[3])
                    # 打印结果
            except:
                print("Error")
                flag = False
                messagebox.showinfo('警告！', '错误！')
            db.close()  # 关闭数据库连接
            if flag:
                messagebox.showinfo('提示', '查询成功！')
                print(self.select_id_id_Check_drug_information)
                self.select_drug_id_gui(self.window)

        else:
            messagebox.showinfo('警告！', '该药品编码不存在,请重新输入')

    def back_drug(self):
        self.window.destroy()  # 回到查询界面
        flag = True

        if str(self.entry_usr_name.get()) in self.id_Check_drug_information:

            db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root',
                                 password='123456')  # 打开数据库连接
            cursor = db.cursor()  # 使用cursor()方法获取操作游标
            sql = "SELECT * FROM medicine WHERE mno = '%s'" % (self.entry_usr_name.get())  # SQL 查询语句
            try:
                # 执行SQL语句
                cursor.execute(sql)
                # 获取所有记录列表
                results = cursor.fetchall()
                for row in results:
                    self.select_id_id_Check_drug_information.append(row[0])
                    self.select_id_name_Check_drug_information.append(row[1])
                    self.select_id_mode_Check_drug_information.append(row[2])
                    self.select_id_mefficacy_Check_drug_information.append(row[3])
                    # 打印结果
            except:
                print("Error")
                flag = False
                messagebox.showinfo('警告！', '错误！')
            db.close()  # 关闭数据库连接
            if flag:
                messagebox.showinfo('提示', '查询成功！')
                print(self.select_id_id_Check_drug_information)
                self.select_drug_id_gui(self.window)

    def drug_name(self):
        self.window = tk.Tk()  # 初始框的声明
        self.window.title('药品名称查询')
        self.window.geometry('400x150')  # 1250x800

        tk.Label(self.window, text='输入名称').place(x=80, y=40)
        self.var_usr_name = tk.StringVar()
        self.entry_usr_name = tk.Entry(self.window, textvariable=self.var_usr_name)
        self.entry_usr_name.place(x=140, y=40)

        # 按钮设置
        btn_login = tk.Button(self.window, text='查询', command=self.select_drug_name)
        btn_login.place(x=100, y=100)
        btn_sign_up = tk.Button(self.window, text='返回', command=self.back_drug)
        btn_sign_up.place(x=250, y=100)

        self.window.mainloop()

    def select_drug_name(self):
        flag = True

        if str(self.entry_usr_name.get()) in self.name_Check_drug_information:

            db = pymysql.connect(host='localhost', port=3306, db='medicine management system', user='root',
                                 password='123456')  # 打开数据库连接
            cursor = db.cursor()  # 使用cursor()方法获取操作游标
            sql = "SELECT * FROM medicine WHERE mname = '%s'" % (self.entry_usr_name.get())  # SQL 查询语句
            try:
                # 执行SQL语句
                cursor.execute(sql)
                # 获取所有记录列表
                results = cursor.fetchall()
                for row in results:
                    self.select_id_id_Check_drug_information.append(row[0])
                    self.select_id_name_Check_drug_information.append(row[1])
                    self.select_id_mode_Check_drug_information.append(row[2])
                    self.select_id_mefficacy_Check_drug_information.append(row[3])
                    # 打印结果
            except:
                print("Error")
                flag = False
                messagebox.showinfo('警告！', '错误！')
            db.close()  # 关闭数据库连接
            if flag:
                messagebox.showinfo('提示', '查询成功！')
                print(self.select_id_id_Check_drug_information)
                self.select_drug_id_gui(self.window)

        else:
            messagebox.showinfo('警告！', '该药品名称不存在,请重新输入')

        

    def select_drug_id_gui(self,parent_window):
        parent_window.destroy()
        self.window_id = tk.Tk()
        self.window_id.title('查询记录')
        self.window_id.geometry('900x360')

        self.frame_center_id = tk.Frame(self.window_id, width=1600, height=300)

        self.columns_id = (
            "编号", "名称", "服用方法", "功效")
        self.tree_id = ttk.Treeview(self.frame_center_id, show="headings", height=30, columns=self.columns_id)  # 18
        self.vbar_id = ttk.Scrollbar(self.frame_center_id, orient=VERTICAL, command=self.tree.yview)
        # 定义树形结构与滚动条
        self.tree_id.configure(yscrollcommand=self.vbar.set)

        # 表格的标题
        self.tree_id.column("编号", width=150, minwidth=100, anchor='center')  # 表示列,不显示
        self.tree_id.column("名称", width=150, minwidth=100, anchor='center')
        self.tree_id.column("服用方法", width=100, minwidth=100, anchor='center')
        self.tree_id.column("功效", width=450, minwidth=100, anchor='center')

        self.tree_id.heading('编号', text="编号", anchor='center')
        self.tree_id.heading("名称", text="名称", anchor='center')
        self.tree_id.heading("服用方法", text="服用方法", anchor='center')
        self.tree_id.heading("功效", text="功效", anchor='center')

        # 调用方法获取表格内容插入\
        self.tree_id.grid(row=0, column=0, sticky=NSEW)
        self.vbar_id.grid(row=0, column=1, sticky=NS)

        # 约束表格不重复插入
        # if str(self.entry_usr_name.get()) not in self.id_id:

        for i in range(min(len(self.select_id_id_Check_drug_information), len(self.select_id_name_Check_drug_information),
                           len(self.select_id_mode_Check_drug_information),
                           len(self.select_id_mefficacy_Check_drug_information))):  # 写入数据
            self.tree_id.insert('', i, values=(self.select_id_id_Check_drug_information[i], self.select_id_name_Check_drug_information[i],
                                            self.select_id_mode_Check_drug_information[i], self.select_id_mefficacy_Check_drug_information[i],
                                            ))


        for col in self.columns_id:  # 绑定函数，使表头可排序
            self.tree_id.heading(col, text=col,
                                 command=lambda _col=col: self.tree_sort_column(self.tree, _col, False))


        self.frame_center_id.grid(row=1, column=0, columnspan=2, padx=4, pady=5)
        self.frame_center_id.grid_propagate(False)


        self.window_id.mainloop()
        self.frame_center_id.tkraise()  # 开始显示主菜单




    def back_select_id(self):
        self.window_id.destroy()

    def select_drug_name_gui(self,parent_window):
        parent_window.destroy()
        self.window_name = tk.Tk()
        self.window_name.title('查询记录')
        self.window_name.geometry('1250x400')

        self.frame_center_id = tk.Frame(self.window_name, width=1600, height=300)

        self.columns_id = (
            "编号", "名称", "服用方法", "功效")
        self.tree_id = ttk.Treeview(self.frame_center_id, show="headings", height=30, columns=self.columns_id)  # 18
        self.vbar_id = ttk.Scrollbar(self.frame_center_id, orient=VERTICAL, command=self.tree.yview)
        # 定义树形结构与滚动条
        self.tree_id.configure(yscrollcommand=self.vbar.set)

        # 表格的标题
        self.tree_id.column("编号", width=150, minwidth=100, anchor='center')  # 表示列,不显示
        self.tree_id.column("名称", width=150, minwidth=100, anchor='center')
        self.tree_id.column("服用方法", width=100, minwidth=100, anchor='center')
        self.tree_id.column("功效", width=450, minwidth=100, anchor='center')

        self.tree_id.heading('编号', text="编号", anchor='center')
        self.tree_id.heading("名称", text="名称", anchor='center')
        self.tree_id.heading("服用方法", text="服用方法", anchor='center')
        self.tree_id.heading("功效", text="功效", anchor='center')

        # 调用方法获取表格内容插入\
        self.tree_id.grid(row=0, column=0, sticky=NSEW)
        self.vbar_id.grid(row=0, column=1, sticky=NS)

        # 约束表格不重复插入
        # if str(self.entry_usr_name.get()) not in self.id_id:

        for i in range(min(len(self.select_id_id_Check_drug_information), len(self.select_id_name_Check_drug_information),
                           len(self.select_id_mode_Check_drug_information),
                           len(self.select_id_mefficacy_Check_drug_information))):  # 写入数据
            self.tree_id.insert('', i, values=(self.select_id_id_Check_drug_information[i], self.select_id_name_Check_drug_information[i],
                                            self.select_id_mode_Check_drug_information[i], self.select_id_mefficacy_Check_drug_information[i],
                                            ))


        for col in self.columns_id:  # 绑定函数，使表头可排序
            self.tree_id.heading(col, text=col,
                                 command=lambda _col=col: self.tree_sort_column(self.tree, _col, False))


        self.frame_center_id.grid(row=1, column=0, columnspan=2, padx=4, pady=5)
        self.frame_center_id.grid_propagate(False)


        self.window_name.mainloop()
        self.frame_center_id.tkraise()  # 开始显示主菜单


    def tree_sort_column(self, tv, col, reverse):  # Treeview、列名、排列方式
        l = [(tv.set(k, col), k) for k in tv.get_children('')]
        l.sort(reverse=reverse)  # 排序方式
        # rearrange items in sorted positions
        for index, (val, k) in enumerate(l):  # 根据排序后索引移动
            tv.move(k, '', index)
        tv.heading(col, command=lambda: self.tree_sort_column(tv, col, not reverse))  # 重写标题，使之成为再点倒序的标题

class Exit_operator:
    def __init__(self,parent_window):
        parent_window.destroy()  # 销毁主界面


if __name__ == '__main__':
    window = tk.Tk()
    StartPage(window)




















